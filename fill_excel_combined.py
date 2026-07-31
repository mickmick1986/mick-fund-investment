# -*- coding: utf-8 -*-
"""
合并写入：D/E/A/C列基础数据 + F/H/I/J/K/L/M/N列(enriched JSON) + O列锚点净值(从_已更新.xlsx读取) + V/W列底谷 + Y列止盈
一次性写入原文件 基金模板（金字塔丛林版）.xlsx
新行自动设置：G/P/Q/R/T/U/X/AA/AB公式 + S(单价=10) + Z(首次建仓="否") + AC(分组="持有")
AD列(基本面状态)：纯展示列，不关联R列。初始值根据回撤自动设置（≥33%→待检查 / <33%→—）
条件格式：用户填写"不通过"→红底醒目，"通过"→绿底，"待检查"→黄底，"—"→灰底
"""
import json
import os
import copy
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ===== 路径配置 =====
_CI_MODE = os.environ.get('CI_MODE', '').lower() == 'true'
if _CI_MODE:
    _REPO_DIR = os.path.dirname(os.path.abspath(__file__))
    ORIGINAL_PATH = os.path.join(_REPO_DIR, '基金模板（金字塔丛林版）.xlsx')
    UPDATED_COPY_PATH = os.path.join(_REPO_DIR, '基金模板（金字塔丛林版）_已更新.xlsx')
    DATA_PATH = os.path.join(_REPO_DIR, 'fund_data_enriched.json')
else:
    ORIGINAL_PATH = r'C:\Users\13697\Desktop\金字塔丛林战法\基金模板（金字塔丛林版）.xlsx'
    UPDATED_COPY_PATH = r'C:\Users\13697\Desktop\金字塔丛林战法\基金模板（金字塔丛林版）_已更新.xlsx'
    DATA_PATH = r'C:\Users\13697\WorkBuddy\2026-06-25-11-50-20\fund_data_enriched.json'
SHEET_NAME = '金字塔丛林补仓'

# M列操作建议字体颜色（与HTML一致）
def get_action_font(signal_text):
    """根据操作建议文字返回带颜色字体"""
    color_map = {
        "强烈补仓": "FF1B5E20",  # 深绿
        "建议补仓": "FF2E7D32",  # 绿色
        "可补仓":   "FFE65100",  # 橙色
        "关注":     "FF5D6D7E",  # 蓝灰
        "观望":     "FF7F8C8D",  # 灰色
        "暂不操作": "FF95A5A6",  # 浅灰
        "警惕回调": "FFC62828",  # 红色
        "考虑止盈": "FF8B0000",  # 暗红（区别于警惕回调的亮红）
    }
    return Font(color=color_map.get(signal_text, "FF000000"), bold=True)


# 清空单元格填充
NO_FILL = PatternFill(fill_type=None)

# ===== AC列分组排序工具 =====
def adjust_formula_row_refs(formula, old_row, new_row):
    """调整公式中的行号引用，把旧行号替换为新行号
    只替换形如 E2、AA2 的单元格引用，不替换纯数字常量"""
    if old_row == new_row:
        return formula
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    # (?<![0-9]) 前面不是数字；(?![0-9]) 后面不是数字，避免把 100 里的 10 当成行号
    pattern = r'(?<![0-9])([A-Z]+)' + str(old_row) + r'(?![0-9])'
    # 用 lambda 避免 re.sub 把 \1 与后面的行号数字混在一起（比如 \1 + "15" → \115 被解析为分组引用）
    return re.sub(pattern, lambda m: m.group(1) + str(new_row), formula)


def reorder_rows_by_group(ws, start_row=2):
    """按AC列分组稳定排序：持有排在前面，关注/其他排在后面。
    保持各组内原有相对顺序不变，即新改成持有的基金会追加到已有持有基金末尾。"""
    max_row = ws.max_row
    if max_row < start_row:
        return []

    # 1. 先完整读取所有数据行到内存（值+样式+公式）
    rows = []
    max_col = max(ws.max_column, 32)
    for row_idx in range(start_row, max_row + 1):
        code = ws.cell(row=row_idx, column=2).value
        group = ws.cell(row=row_idx, column=29).value  # AC列

        cells = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cells.append({
                'value': cell.value,
                'number_format': cell.number_format,
                'font': copy.copy(cell.font),
                'fill': copy.copy(cell.fill),
                'alignment': copy.copy(cell.alignment),
                'border': copy.copy(cell.border),
            })

        rows.append({
            'orig_row': row_idx,
            'code': str(code).strip() if code else '',
            'group': str(group).strip() if group else '',
            'cells': cells,
        })

    # 2. 稳定排序：持有=0，其他=1
    rows.sort(key=lambda r: 0 if r['group'] == '持有' else 1)

    # 3. 按排序后的顺序写回工作表
    for new_idx, row_data in enumerate(rows):
        new_row = start_row + new_idx
        old_row = row_data['orig_row']
        for col_idx, cell_data in enumerate(row_data['cells'], 1):
            cell = ws.cell(row=new_row, column=col_idx)
            val = cell_data['value']
            if isinstance(val, str) and val.startswith('='):
                val = adjust_formula_row_refs(val, old_row, new_row)
            cell.value = val
            cell.number_format = cell_data['number_format']
            cell.font = cell_data['font']
            cell.fill = cell_data['fill']
            cell.alignment = cell_data['alignment']
            cell.border = cell_data['border']

    # 4. 清空尾部多余的行（原来位于中间的空行被排到最下面后需要清空）
    for row_idx in range(start_row + len(rows), max_row + 1):
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    return rows


# ===== AD列（基本面状态）纯展示配置 =====
AD_COL = 30  # AD列号
AD_THRESHOLD = -0.33  # 回撤≥33%触发"待检查"

AD_HEADER_FONT = Font(bold=True, size=11, color="FFFFFFFF")
AD_HEADER_FILL = PatternFill(start_color="FF2E7D32", end_color="FF2E7D32", fill_type="solid")
AD_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

AD_PENDING_FONT = Font(bold=True, color="FFF57F17")
AD_PENDING_FILL = PatternFill(start_color="FFFFF9C4", end_color="FFFFF9C4", fill_type="solid")
AD_DASH_FONT = Font(color="FF9E9E9E")
AD_DASH_FILL = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")

AD_RED_FILL = PatternFill(start_color="FFC62828", end_color="FFC62828", fill_type="solid")
AD_RED_FONT = Font(bold=True, color="FFFFFFFF", size=11)
AD_GREEN_FILL = PatternFill(start_color="FF1B5E20", end_color="FF1B5E20", fill_type="solid")
AD_GREEN_FONT = Font(bold=True, color="FFFFFFFF", size=11)


def setup_ad_column(ws, max_row):
    """确保AD列表头存在，并添加条件格式（幂等，可安全重复调用）"""
    from openpyxl.formatting.rule import FormulaRule

    header = ws.cell(row=1, column=AD_COL).value
    if not header or header != "基本面状态":
        ws.cell(row=1, column=AD_COL).value = "基本面状态"
        ws.cell(row=1, column=AD_COL).font = AD_HEADER_FONT
        ws.cell(row=1, column=AD_COL).fill = AD_HEADER_FILL
        ws.cell(row=1, column=AD_COL).alignment = AD_CENTER

    if ws.column_dimensions['AD'].width is None or ws.column_dimensions['AD'].width < 28:
        ws.column_dimensions['AD'].width = 28

    cf_range = f"AD2:AD{max(max_row, 30)}"
    to_remove = []
    for cf_range_key in list(ws.conditional_formatting._cf_rules.keys()):
        if 'AD' in str(cf_range_key):
            to_remove.append(cf_range_key)
    for key in to_remove:
        del ws.conditional_formatting._cf_rules[key]

    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("不通过",AD2))'],
        fill=AD_RED_FILL, font=AD_RED_FONT, stopIfTrue=True
    ))
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f'AND(ISNUMBER(SEARCH("通过",AD2)),NOT(ISNUMBER(SEARCH("不通过",AD2))))'],
        fill=AD_GREEN_FILL, font=AD_GREEN_FONT, stopIfTrue=True
    ))
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f'AD2="待检查"'],
        fill=AD_PENDING_FILL, font=AD_PENDING_FONT, stopIfTrue=True
    ))
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f'AD2="—"'],
        fill=AD_DASH_FILL, font=AD_DASH_FONT, stopIfTrue=True
    ))


def set_ad_cell(ws, row, dd):
    """设置AD列初始值（仅当AD为空或"—"时更新，不覆盖用户手动填写的值）"""
    ad_cell = ws.cell(row=row, column=AD_COL)
    existing = ad_cell.value
    if existing and str(existing).strip() not in ("", "—"):
        return
    ad_cell.alignment = AD_CENTER
    if dd <= AD_THRESHOLD:
        ad_cell.value = "待检查"
        ad_cell.font = AD_PENDING_FONT
        ad_cell.fill = AD_PENDING_FILL
    else:
        ad_cell.value = "—"
        ad_cell.font = AD_DASH_FONT
        ad_cell.fill = AD_DASH_FILL


def setup_new_row_formulas(ws, row):
    """为新行设置所有公式和默认值（仅在该行G列无公式时触发）"""
    g_val = ws.cell(row=row, column=7).value
    if g_val and str(g_val).startswith('='):
        return False

    ws.cell(row=row, column=7).value = f'=(E{row}-D{row})/D{row}'
    ws.cell(row=row, column=7).number_format = '0.00%'
    ws.cell(row=row, column=16).value = f'=(E{row}-O{row})/O{row}'
    ws.cell(row=row, column=16).number_format = '0.00%'
    ws.cell(row=row, column=17).value = f'=P{row}'
    ws.cell(row=row, column=17).number_format = '0.00%'
    ws.cell(row=row, column=18).value = f'=IF(OR(L{row}="强烈补仓",L{row}="建议补仓",L{row}="可补仓"),MAX(IF(Q{row}<0,INT(-Q{row}*100),0),1),0)'
    if not ws.cell(row=row, column=19).value:
        ws.cell(row=row, column=19).value = 10
    ws.cell(row=row, column=20).value = f'=N{row}*R{row}'
    ws.cell(row=row, column=21).value = f'=S{row}*T{row}'
    ws.cell(row=row, column=24).value = f'=(E{row}-W{row})/W{row}'
    ws.cell(row=row, column=24).number_format = '0.00%'
    if not ws.cell(row=row, column=26).value:
        ws.cell(row=row, column=26).value = "否"
    ws.cell(row=row, column=27).value = f'=IF(Z{row}="是",IF(G{row}<0,INT(-G{row}*100),"0")*N{row},"")'
    ws.cell(row=row, column=28).value = f'=AA{row}*S{row}'
    if not ws.cell(row=row, column=29).value:
        ws.cell(row=row, column=29).value = "持有"
    return True


def format_valuation(fund):
    """格式化K列 PE/PB估值"""
    signal = fund['val_signal']
    if signal and (signal.startswith('PE=') or signal.startswith('PB=')):
        return signal
    if fund['val_metric'] == '不适用':
        return signal
    metric = fund['val_metric']
    value = fund['val_value']
    if value is None:
        return signal
    return f"{metric}={value:.2f} {signal}"


def read_anchor_navs():
    """从_已更新.xlsx读取O列(第15列)锚点净值；若不存在则从原文件读取保留"""
    paths_to_try = [UPDATED_COPY_PATH, ORIGINAL_PATH]
    for path in paths_to_try:
        if not os.path.exists(path):
            continue
        wb = load_workbook(path, data_only=True)
        ws = wb[SHEET_NAME]
        anchor_map = {}
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row=row, column=2).value
            nav = ws.cell(row=row, column=15).value
            if code and nav is not None:
                code = str(code).strip()
                anchor_map[code] = nav
        wb.close()
        if anchor_map:
            print(f"  O列读取: 从 {path} 读取 {len(anchor_map)} 只基金锚点净值")
            return anchor_map
    print(f"⚠️ 未找到O列锚点净值数据，跳过O列写入")
    return {}


def main():
    # 1. 读取enriched JSON数据
    with open(DATA_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    funds = data['funds']
    final_nav_mode = data.get('data_mode') == 'final_nav'
    fund_by_code = {f['code']: f for f in funds}
    mode_name = '早间确认净值模式' if final_nav_mode else '盘中估值模式'
    print(f"📊 加载了 {len(funds)} 只基金数据(enriched JSON) | {mode_name}")

    # 2. 从_已更新.xlsx读取O列锚点净值
    print(f"\n📖 读取_已更新.xlsx中的O列锚点净值...")
    anchor_navs = read_anchor_navs()
    print(f"  共读取 {len(anchor_navs)} 只基金的锚点净值")

    # 3. 打开原文件
    print(f"\n📁 打开原文件: {ORIGINAL_PATH}")
    wb = load_workbook(ORIGINAL_PATH)
    ws = wb[SHEET_NAME]

    # ══════════════════════════════════════════════════════════════
    # 【AC列分组排序】：持有排在前面，关注排在后面
    # 稳定排序，保持各组内原有相对顺序不变
    # ══════════════════════════════════════════════════════════════
    print(f"\n📋 按AC列分组排序（持有→关注）...")
    sorted_rows = reorder_rows_by_group(ws, start_row=2)
    if sorted_rows:
        hold_count = sum(1 for r in sorted_rows if r['group'] == '持有')
        watch_count = len(sorted_rows) - hold_count
        print(f"  ✓ 排序完成：持有 {hold_count} 只，关注/其他 {watch_count} 只")

    # ══════════════════════════════════════════════════════════════
    # 【列宽保护锁】：快照用户手动调整的列宽，保存前强制恢复
    # 用户是列宽的唯一主人，任何代码都不得修改列宽
    # 注意：必须用 ws.column_dimensions[col] 直接访问，不能用 .get()
    # 因为 openpyxl 的 DimensionHolder.__getitem__ 会通过 default_factory
    # 读取 WPS/Excel 存储的实际列宽，而 .get() 对不在字典中的列返回 None
    # ══════════════════════════════════════════════════════════════
    from openpyxl.utils import get_column_letter
    col_width_snapshot = {}
    max_col = max(ws.max_column, 32)  # 至少扫描到AF列
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        # 直接访问触发 default_factory，确保读到 WPS 存储的真实列宽
        dim = ws.column_dimensions[col_letter]
        if dim.width is not None:
            w = dim.width
            # M列(倍投等级)定死宽度8
            if col_letter == 'M':
                w = 8
            # N列(补仓倍数)硬上限：表头"补仓倍数"4字+数据1.5x等，超过7.8强制截断
            elif col_letter == 'N' and w > 7.8:
                w = 7.8
            col_width_snapshot[col_letter] = w
    print(f"🔒 已快照 {len(col_width_snapshot)} 列宽，保存时将原样恢复")

    # AD列(基本面状态)表头 + 条件格式（幂等，可安全重复调用）
    setup_ad_column(ws, ws.max_row)
    print(f"📋 AD列(基本面状态)已就绪：表头+条件格式（纯展示，不关联R列）")

    filled_k = 0
    filled_o = 0
    new_rows = 0
    not_found = []

    # 4. 遍历Excel每一行，写入全部列
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row=row, column=2).value
        if not code:
            continue
        code = str(code).strip()

        # --- 写入基础列+指标列 (来自enriched JSON) ---
        if code in fund_by_code:
            fund = fund_by_code[code]

            # === 新行检测：设置公式和默认值 ===
            is_new = setup_new_row_formulas(ws, row)
            if is_new:
                new_rows += 1
                print(f"  🆕 新行{row}: {code} 已设置公式和默认值")

            # === A列: 类型（仅空行填充）===
            if not ws.cell(row=row, column=1).value:
                ws.cell(row=row, column=1).value = fund.get('category', '')

            # === C列: 基金名（仅空行填充）===
            if not ws.cell(row=row, column=3).value:
                ws.cell(row=row, column=3).value = fund.get('name', code)

            # === D列: 历史最高净值（每次更新）===
            ws.cell(row=row, column=4).value = fund['all_time_high']
            ws.cell(row=row, column=4).number_format = '0.0000'

            # === E列: 最新净值（每次更新）===
            ws.cell(row=row, column=5).value = fund['latest_nav']
            ws.cell(row=row, column=5).number_format = '0.0000'

            # 最终结算模式：F列清空，且G/Q/X只按确认净值计算，防止F与E重复计入当日涨跌。
            if final_nav_mode:
                ws.cell(row=row, column=6).value = None
            else:
                dc = fund['daily_change']
                ws.cell(row=row, column=6).value = None if dc is None else dc / 100
            ws.cell(row=row, column=6).number_format = '0.00%'
            ws.cell(row=row, column=7).value = f'=(E{row}-D{row})/D{row}'
            ws.cell(row=row, column=7).number_format = '0.00%'
            ws.cell(row=row, column=16).value = f'=(E{row}-O{row})/O{row}'
            ws.cell(row=row, column=16).number_format = '0.00%'
            ws.cell(row=row, column=17).value = f'=P{row}'
            ws.cell(row=row, column=17).number_format = '0.00%'
            ws.cell(row=row, column=24).value = f'=(E{row}-W{row})/W{row}'
            ws.cell(row=row, column=24).number_format = '0.00%'

            # H列: RSI数值
            ws.cell(row=row, column=8).value = fund['rsi']
            ws.cell(row=row, column=8).number_format = '0.0'
            
            # I列: RSI信号
            ws.cell(row=row, column=9).value = fund['rsi_signal']
            
            # J列: PE/PB估值
            ws.cell(row=row, column=10).value = format_valuation(fund)
            
            # K列: 综合信号（含义文字）
            ws.cell(row=row, column=11).value = fund['comp_meaning']
            ws.cell(row=row, column=11).fill = NO_FILL
            
            # L列: 操作建议（信号文字 + 字体颜色）
            ws.cell(row=row, column=12).value = fund['comp_signal']
            ws.cell(row=row, column=12).font = get_action_font(fund['comp_signal'])
            ws.cell(row=row, column=12).fill = NO_FILL
            
            # M列: 倍投等级
            ws.cell(row=row, column=13).value = fund['level']
            
            # N列: 补仓倍数
            ws.cell(row=row, column=14).value = fund['multiplier']

            # V列: 最近底谷日期（每次更新，随主流程刷新）
            if fund.get('trough_date'):
                ws.cell(row=row, column=22).value = fund['trough_date']

            # W列: 最近底谷净值（每次更新，随主流程刷新）
            if fund.get('trough_nav') is not None:
                ws.cell(row=row, column=23).value = fund['trough_nav']
                ws.cell(row=row, column=23).number_format = '0.0000'

            # Y列: 止盈建议（止盈等级·卖XX%）
            tp_display = fund.get('tp_display')
            if tp_display:
                ws.cell(row=row, column=25).value = tp_display
                ws.cell(row=row, column=25).font = Font(color="FF8B0000", bold=True)
            else:
                ws.cell(row=row, column=25).value = None
                ws.cell(row=row, column=25).font = Font()

            # AD列: 基本面状态（纯展示，不关联R列）
            dd_raw = (fund['latest_nav'] - fund['all_time_high']) / fund['all_time_high']
            set_ad_cell(ws, row, dd_raw)

            filled_k += 1
            dc_str = "休市" if fund['daily_change'] is None else f"{fund['daily_change']:+.2f}%"
            print(f"  ✓ [{code}] {fund['name']}: 净值{fund['latest_nav']:.4f}, 涨跌{dc_str}, 等级{fund['level']}")
        else:
            not_found.append((row, code))

        # --- 写入O列: 锚点净值 (来自_已更新.xlsx) ---
        if code in anchor_navs:
            ws.cell(row=row, column=15).value = anchor_navs[code]
            ws.cell(row=row, column=15).number_format = '0.0000'
            filled_o += 1
            print(f"  ✓ O列[{code}] 锚点净值: {anchor_navs[code]}")

    # 5. 【列宽保护锁】：保存前强制恢复用户手动调整的列宽
    for col_letter, width in col_width_snapshot.items():
        ws.column_dimensions[col_letter].width = width

    # 6. 保存到原文件
    wb.save(ORIGINAL_PATH)
    wb.close()
    print(f"\n{'='*60}")
    print(f"✅ 完成！（列宽已原样恢复 🔒）")
    print(f"  指标数据: {filled_k} 行")
    print(f"  O列锚点净值: {filled_o} 行")
    if new_rows:
        print(f"  🆕 新行: {new_rows} 行（已自动设置公式+默认值）")
    if not_found:
        print(f"  ⚠️ 未匹配JSON（需先运行fetch+enrich）: {not_found}")
    print(f"  📄 已保存到原文件: {ORIGINAL_PATH}")


if __name__ == '__main__':
    main()
