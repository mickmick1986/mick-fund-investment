#!/usr/bin/env python3
"""
金字塔丛林补仓表 - 数据增强脚本 V2.1
V2重大改进：
  1. 集成蛋卷 /djapi/index_eva/dj API (63指数 PE+PB+pb_flag+分位%)
  2. 按基金类型智能选择PE或PB (光伏/有色/畜牧/芯片→PB, 券商/军工pb_flag→PB, 科创50 PE>200→PB)
  3. 腾讯财经API作为PE/PB备选数据源
  4. 具体N/A原因（不再使用通用"数据暂缺"）
  5. VALUATION_REF定制阈值评低估/适中/偏高/高估
  6. F列当天涨跌幅：新浪基金估值接口 fu_{code} 覆盖全部25只基金（含黄金/主动/电网），指数/ETF/lsjz作为回退

V2.1新增改进：
  7. 中证官网perf API获取PE（当天更新，YYYYMMDD日期格式）→ 类型二7只指数PE全覆盖
  8. 乐咕乐股legulegu API获取PB（cookie/csrf缓存）→ 光伏等指数PB数据
  9. PB首选不可用时PE作为备选（芯片/畜牧标注说明）
  10. 指数代码修正：025832电网→931994，024194卫星→980018
  11. 估值链路：蛋卷→腾讯→中证perf API(PE)→legulegu(PB)
"""
import json
import os
import re
import requests
import time
import subprocess
import hashlib
import urllib.request
import io
import argparse
from datetime import datetime, timedelta, date
from hashlib import md5
from bs4 import BeautifulSoup
from openpyxl import load_workbook

UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
API_HEADERS = {'User-Agent': UA, 'Referer': 'https://fundf10.eastmoney.com/'}

# ===== 配置 =====
BASE = 10  # 补仓基准金额

# ===== 9级倍投等级表（极限区分档：48-55%→5x, 55-65%→6x, >65%→7x）=====
LEVEL_TABLE = [
    ("停止区", 0, 13, 0),
    ("观望区", 13, 15, 0.5),
    ("倍投1级", 15, 20, 1),
    ("倍投2级", 20, 25, 1.5),
    ("倍投3级", 25, 33, 2),
    ("倍投4级", 33, 40, 3),
    ("倍投5级", 40, 48, 4),
    ("极限1级", 48, 55, 5),
    ("极限2级", 55, 65, 6),
    ("极限3级", 65, 100, 7),
]

# ===== 止盈等级表（涨幅 vs 止盈锚点）=====
# sell_pct 已转换为整数百分比（无小数）
TAKE_PROFIT_LEVELS = [
    # (name, min_gain%, max_gain%, sell_pct_int)
    ("观察区",  0,   5,  0),   # 不卖
    ("止盈1级", 5,  10, 17),   # 原1/6 ≈ 16.7% → 17%
    ("止盈2级", 10, 15, 25),   # 原1/4 = 25%
    ("止盈3级", 15, 25, 33),   # 原1/3 ≈ 33.3% → 33%
    ("止盈4级", 25, 40, 50),   # 原1/2 = 50%
    ("清仓区",  40, 999, 100), # 全部
]

# ===== 市场状态自适应止盈参数表 =====
# 不同市场状态下止盈策略的参数调整（Regime Switching）
MARKET_REGIME_TABLE = {
    "early_bull": {
        "name": "牛市初期（熊转牛）",
        "desc": "市场刚从熊市反转，上涨空间大，避免卖飞",
        "rsi_tp_strong": 80,     # RSI>80 才无条件触发（默认75→提高）
        "rsi_tp_val": 75,       # RSI>75 + 估值偏高 才触发（默认70→提高）
        "rsi_tp_x_gain": 75,    # RSI>75 + X列涨幅>25% 才触发
        "sell_pct_scale": 0.6,  # 卖出比例×0.6（极度耐心，少卖多留）
        "anchor_clear_rsi": 45, # RSI回落到45以下才清除锚点
    },
    "mid_bull": {
        "name": "牛市中期",
        "desc": "趋势确立，平衡止盈与持有",
        "rsi_tp_strong": 75,     # 当前默认值
        "rsi_tp_val": 70,
        "rsi_tp_x_gain": 70,
        "sell_pct_scale": 1.0,  # 正常比例
        "anchor_clear_rsi": 50,
    },
    "late_bull": {
        "name": "牛市后期",
        "desc": "估值偏高，随时可能回调，落袋为安",
        "rsi_tp_strong": 65,     # RSI>65 就触发（降低门槛）
        "rsi_tp_val": 60,       # RSI>60 + 估值偏高
        "rsi_tp_x_gain": 60,
        "sell_pct_scale": 1.5,  # 卖出比例×1.5（积极止盈）
        "anchor_clear_rsi": 55,
    },
    "correction": {
        "name": "调整期",
        "desc": "市场回调中，止盈暂停，专注补仓",
        "rsi_tp_strong": 999,    # 实际不触发
        "rsi_tp_val": 999,
        "rsi_tp_x_gain": 999,
        "sell_pct_scale": 0,    # 不卖出
        "anchor_clear_rsi": 40,
    },
}

# 止盈锚点持久化文件
TP_STATE_FILE = 'take_profit_state.json'

# CI模式检测
_CI_MODE = os.environ.get('CI_MODE', '').lower() == 'true'

# Excel模板路径（读取V/W底谷数据）
if _CI_MODE:
    EXCEL_TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '基金模板（金字塔丛林版）.xlsx')
else:
    EXCEL_TEMPLATE_PATH = r'C:\Users\13697\Desktop\金字塔丛林战法\基金模板（金字塔丛林版）.xlsx'

# ===== 基金→跟踪指数映射 =====
INDEX_MAP = {
    "000218": None,       # 黄金，无指数
    "024194": "980018",   # 国证卫星通信（原932096不存在，正确为980018国证指数）
    "024202": None,       # 主动基金
    "001475": "399967",   # 中证军工
    "013477": "930986",   # 金融科技
    "005453": "399386",   # 1000医药
    "008928": "000932",   # 中证消费
    "003095": "399989",   # 中证医疗
    "011102": "931151",   # 中证光伏
    "003984": "399808",   # 中证新能
    "010769": "000949",   # 中证农业
    "025832": "931994",   # 中证电网设备主题指数（原930700为上海改革发展指数，错误）
    "004432": "000819",   # 有色金属
    "008887": "990001",   # 国证芯片
    "018344": "H30590",   # 机器人
    "001790": "930721",   # CS智汽车
    "014414": "930707",   # 中证畜牧
    "006098": "399975",   # 证券公司
    "011608": "000688",   # 科创50
    "003765": "399006",   # 创业板
    "025656": "000905",   # 中证500
    "025880": "000300",   # 沪深300
    "021420": "399678",   # 信创
    "013402": "HSTECH",   # 恒生科技
    "161725": "399997",   # 中证白酒
}

# ===== 估值阈值参考表 =====
VALUATION_REF = {
    "399967": {"pe": (45, 60, 80), "pb": (2.0, 3.5, 5.5), "desc": "中证军工"},  # pb_flag=True
    "930986": {"pe": (30, 45, 65), "desc": "金融科技"},
    "399989": {"pe": (25, 40, 60), "desc": "中证医疗"},
    "931151": {"pb": (1.2, 2.0, 3.5), "desc": "中证光伏"},
    "399808": {"pe": (20, 35, 55), "desc": "中证新能"},
    "000949": {"pe": (20, 30, 45), "desc": "中证农业"},
    "931994": {"pe": (20, 35, 55), "desc": "电网设备"},
    "000819": {"pb": (1.5, 2.5, 4.0), "desc": "有色金属"},
    "990001": {"pb": (2.0, 3.5, 6.0), "pe": (40, 70, 120), "desc": "国证芯片"},  # PB首选，PE备用
    "930721": {"pe": (25, 40, 60), "desc": "CS智汽车"},
    "930707": {"pb": (1.5, 2.5, 4.0), "pe": (15, 25, 40), "desc": "中证畜牧"},  # PB首选，PE备用
    "399975": {"pe": (15, 22, 35), "pb": (1.0, 1.5, 2.5), "desc": "证券公司"},  # pb_flag=True
    "000688": {"pe": (35, 55, 80), "pb": (3.0, 5.0, 8.0), "desc": "科创50"},  # PE>200→PB
    "399006": {"pe": (30, 45, 65), "desc": "创业板指"},
    "000905": {"pe": (20, 28, 40), "desc": "中证500"},
    "000300": {"pe": (11, 13, 16), "desc": "沪深300"},
    "399678": {"pe": (35, 50, 70), "desc": "深证信创"},
    "399997": {"pe": (18, 28, 45), "pb": (3.0, 5.0, 8.0), "desc": "中证白酒"},
    "000932": {"pe": (18, 25, 35), "desc": "中证消费"},
    "399386": {"pe": (25, 35, 50), "desc": "1000医药"},
    "980018": {"pb": (3.0, 5.0, 8.0), "pe": (30, 50, 80), "desc": "国证卫星通信"},  # 军工PB首选，PE备用
    "HSTECH": {"pe": (15, 25, 40), "pb": (1.5, 2.5, 4.0), "desc": "恒生科技"},
}

# ===== 我们的指数 → 蛋卷API代码映射 =====
DANJUAN_CODE_MAP = {
    "399967": "SZ399967",   # 中证军工
    "000932": "SH000932",   # 主要消费
    "399989": "SZ399989",   # 中证医疗
    "399975": "SZ399975",   # 证券公司
    "000688": "SH000688",   # 科创50
    "399006": "SZ399006",   # 创业板
    "000905": "SH000905",   # 中证500
    "000300": "SH000300",   # 沪深300
    "399997": "SZ399997",   # 中证白酒
    "HSTECH": "HKHSTECH",   # 恒生科技
}

# ===== 新浪实时指数/ETF行情映射 =====
# 指数 → 新浪代码（标准A股指数格式）
SINA_CODE_MAP = {
    "000300": "sh000300",   # 沪深300
    "000905": "sh000905",   # 中证500
    "399997": "sz399997",   # 中证白酒
    "399967": "sz399967",   # 中证军工
    "399989": "sz399989",   # 中证医疗
    "399975": "sz399975",   # 证券公司
    "399006": "sz399006",   # 创业板
    "000688": "sh000688",   # 科创50
    "000932": "sh000932",   # 中证消费
    "000819": "sh000819",   # 有色金属
    "399808": "sz399808",   # 中证新能
    "000949": "sh000949",   # 中证农业
    "399386": "sz399386",   # 1000医药
    "399678": "sz399678",   # 深证信创
    "HSTECH": "hkHSTECH",   # 恒生科技
}

# CSI自定义指数 → 跟踪ETF的新浪代码（指数本身无实时行情，用ETF替代）
ETF_CODE_MAP = {
    "931151": "sz159857",   # 光伏ETF → 中证光伏
    "990001": "sz159995",   # 芯片ETF → 国证芯片
    "930707": "sz159865",   # 畜牧ETF → 中证畜牧
    "980018": "sz159811",   # 5G通信ETF → 国证卫星通信（近似）
    "930986": "sz159851",   # 金融科技ETF
    "H30590": "sz159770",   # 机器人ETF
    "930721": "sz159889",   # 智车ETF → CS智汽车
    "000949": "sz159825",   # 农业ETF → 中证农业（指数000949已停更）
}


def fetch_danjuan_index_valuation():
    """获取蛋卷63指数估值数据"""
    try:
        url = 'https://danjuanfunds.com/djapi/index_eva/dj'
        r = requests.get(url, headers={'User-Agent': UA}, timeout=15)
        data = r.json()
        items = data.get('data', {}).get('items', [])
        result = {}
        for item in items:
            code = item.get('index_code', '')
            result[code] = {
                'name': item.get('name', ''),
                'pe': item.get('pe'),
                'pb': item.get('pb'),
                'pb_flag': item.get('pb_flag', False),
                'pe_percentile': item.get('pe_percentile'),
                'pb_percentile': item.get('pb_percentile'),
                'roe': item.get('roe'),
            }
        return result
    except Exception as e:
        print(f"  [WARN] 蛋卷API获取失败: {e}")
        return {}


def get_index_pe_pb_tencent(index_code):
    """从腾讯财经获取指数PE/PB (备选)"""
    try:
        prefix = "sh" if index_code.startswith(("0", "6", "9")) else "sz"
        # 特殊处理 H 开头的指数
        if index_code.startswith("H"):
            prefix = "sz"
            code = index_code[1:]
        else:
            code = index_code
        url = f"https://qt.gtimg.cn/q={prefix}{code}"
        req = urllib.request.Request(url)
        req.add_header("User-Agent", UA)
        resp = urllib.request.urlopen(req, timeout=10)
        text = resp.read().decode("gbk")
        parts = text.split('"')
        if len(parts) < 3 or not parts[1]:
            return None
        vals = parts[1].split("~")
        if len(vals) < 47:
            return None
        pe = 0
        pb = 0
        for idx in [39, 38, 37, 40]:
            try:
                v = float(vals[idx])
                if 0 < v < 500:
                    pe = v
                    break
            except:
                pass
        for idx in [46, 45, 47, 44]:
            try:
                v = float(vals[idx])
                if 0 < v < 20:
                    pb = v
                    break
            except:
                pass
        return {"pe": pe, "pb": pb}
    except:
        return None


# ===== 中证官网 perf API: 获取指数PE(TTM) =====
def fetch_csindex_pe_perf(index_code):
    """
    中证官网perf API获取指数PE（当天更新，YYYYMMDD日期格式）
    覆盖7/8只类型二指数（980018国证指数除外）
    返回: {"pe": float, "date": str, "close": float, "source": "中证官网"}
    """
    if index_code is None or index_code == "HSTECH":
        return None
    url = "https://www.csindex.com.cn/csindex-home/perf/index-perf"
    end_date = datetime.now().strftime("%Y%m%d")
    start_date = (datetime.now() - timedelta(days=180)).strftime("%Y%m%d")
    params = {"indexCode": index_code, "startDate": start_date, "endDate": end_date}
    headers = {"User-Agent": UA, "Referer": "https://www.csindex.com.cn/"}
    try:
        r = requests.get(url, params=params, headers=headers, timeout=15)
        data = r.json()
        if "data" in data and len(data.get("data", [])) > 0:
            latest = data["data"][-1]
            peg = latest.get("peg")
            if peg is not None and float(peg) > 0:
                return {
                    "pe": float(peg),
                    "date": latest.get("tradeDate", ""),
                    "close": float(latest.get("close", 0)) if latest.get("close") else None,
                    "source": "中证官网"
                }
    except Exception as e:
        print(f"  [perf API] {index_code}: {e}")
    return None


# ===== 乐咕乐股 legulegu API: 获取指数PB =====
_legulegu_cookies = None
_legulegu_csrf = None
_legulegu_ts = 0

def _get_legulegu_auth():
    """获取legulegu cookie和csrf-token（缓存，避免重复请求）"""
    global _legulegu_cookies, _legulegu_csrf, _legulegu_ts
    # 缓存5分钟
    if _legulegu_cookies and (time.time() - _legulegu_ts) < 300:
        return _legulegu_cookies, _legulegu_csrf

    ref_url = "https://legulegu.com/stockdata/zz500-ttm-lyr"
    session = requests.Session()
    session.headers.update({
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "User-Agent": UA,
        "Referer": "https://legulegu.com/",
    })
    try:
        r = session.get(ref_url, timeout=15)
        soup = BeautifulSoup(r.text, features="lxml")
        csrf_tag = soup.find(name="meta", attrs={"name": "_csrf"})
        if csrf_tag:
            _legulegu_cookies = session.cookies
            _legulegu_csrf = csrf_tag.attrs["content"]
            _legulegu_ts = time.time()
            return _legulegu_cookies, _legulegu_csrf
    except Exception as e:
        print(f"  [legulegu] csrf获取失败: {e}")
    return None, None


def fetch_legulegu_pb(index_code):
    """
    乐咕乐股legulegu API获取指数PB（当天更新）
    覆盖legulegu数据库中的指数（931151验证可用）
    返回: {"pb": float, "addPb": float, "date": str, "source": "乐咕乐股"}
    """
    cookies, csrf = _get_legulegu_auth()
    if not cookies or not csrf:
        return None

    token = md5(datetime.now().date().isoformat().encode("utf-8")).hexdigest()
    url = "https://legulegu.com/api/stockdata/index-basic-pb"
    headers = {"User-Agent": UA, "X-CSRF-Token": csrf, "Referer": "https://legulegu.com/"}

    # 尝试不同后缀: .CSI / .SH / .SZ
    for suffix in [".CSI", ".SH", ".SZ"]:
        params = {"token": token, "indexCode": f"{index_code}{suffix}"}
        try:
            r = requests.get(url, params=params, cookies=cookies, headers=headers, timeout=15)
            data = r.json()
            if "data" in data and data.get("data") and len(data["data"]) > 0:
                latest = data["data"][-1]
                pb = latest.get("pb")
                if pb is not None and float(pb) > 0:
                    return {
                        "pb": float(pb),
                        "addPb": float(latest.get("addPb", 0)) if latest.get("addPb") else None,
                        "middlePb": float(latest.get("middlePb", 0)) if latest.get("middlePb") else None,
                        "date": latest.get("date", ""),
                        "source": "乐咕乐股"
                    }
        except Exception:
            pass
    return None


# ===== 市场状态检测模块 =====
def calc_rsi_from_closes(closes, period=14):
    """从收盘价序列计算RSI（closes: [旧→新]，时间顺序）"""
    if len(closes) < period + 1:
        return None
    changes = [closes[i+1] - closes[i] for i in range(len(closes)-1)]
    recent = changes[-(period):]
    gains = sum(c for c in recent if c > 0) / period
    losses = sum(-c for c in recent if c < 0) / period
    if losses == 0:
        return 100.0
    rs = gains / losses
    return round(100 - 100 / (1 + rs), 1)


# ===== 韭圈儿恐贪指数解密脚本路径 =====
if _CI_MODE:
    _JQ_DECRYPT_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'decrypt_helper.js')
    _JQ_NODE_BIN = 'node'
    _JQ_NODE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'node_modules')
else:
    _JQ_DECRYPT_SCRIPT = r'C:\Users\13697\WorkBuddy\2026-06-20-18-01-34\decrypt_helper.js'
    _JQ_NODE_BIN = r'C:\Users\13697\.workbuddy\binaries\node\versions\22.22.2\node.exe'
    _JQ_NODE_PATH = r'C:\Users\13697\.workbuddy\binaries\node\workspace\node_modules'
_JQ_SECRET_KEY = 'flkjer45#f54trg2d1r54'
_JQ_TEMP_ENCRYPTED = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_jq_temp_encrypted.txt')
_JQ_TEMP_DECRYPTED = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_jq_temp_decrypted.json')


def _sign_jiucaishuo(data):
    """韭圈儿 fadrefa 签名算法"""
    data['type'] = 'pc'
    data['version'] = '2.2.7'
    if 'authtoken' not in data:
        data['authtoken'] = ''
    data['act_time'] = int(time.time() * 1000)

    sorted_keys = sorted(data.keys())
    o = ''
    for r in sorted_keys:
        val = data[r]
        if val is None or val == '':
            data[r] = ''
            continue
        if r == 'rid' and isinstance(val, (int, float)):
            val = str(val)
            data[r] = val
        if val == 0 or val == '0':
            o += '0'
            continue
        if val and not isinstance(val, (dict, list)):
            o += str(val)
    o += _JQ_SECRET_KEY
    u = hashlib.md5(o.encode('utf-8')).hexdigest()

    params = {
        'd': u[2:4], 'f': u[5:6], 'V': u[24:26], 'U': u[12:14], 'S': u[25:26],
        'R': u[16:19], 'L': u[16:17], 'c': u[29:31], 'h': u[26:27], 'm': u[6:8],
        'v': u[1:2], 'N': u[21:23], 'y': u[0:2], 'K': u[21:23], 'D': u[14:16],
        '$': u[29:32], 'x': u[30:31], 'A': u[9:11], 'C': u[27:29], 'T': u[17:19],
        'B': u[18:19], 'k': u[6:8], 'j': u[11:14], 'I': u[26:27], 'F': u[17:21],
        'E': u[23:25], 'H': u[31:32], 'O': u[25:27], 'w': u[8:9], 'P': u[11:12],
        'z': u[2:5], 'q': u[9:11],
    }
    data.update({
        'tirgkjfs': params['y'], 'abiokytke': params['N'], 'u54rg5d': params['d'],
        'kf54ge7': params['q'], 'tiklsktr4': params['v'], 'lksytkjh': params['F'],
        'sbnoywr': params['E'], 'bgd7h8tyu54': params['w'], 'y654b5fs3tr': params['O'],
        'bioduytlw': params['f'], 'bd4uy742': params['I'], 'h67456y': params['R'],
        'bvytikwqjk': params['m'], 'ngd4uy551': params['T'], 'bgiuytkw': params['A'],
        'nd354uy4752': params['x'], 'ghtoiutkmlg': params['j'], 'bd24y6421f': params['V'],
        'tbvdiuytk': params['L'], 'ibvytiqjek': params['D'], 'jnhf8u5231': params['C'],
        'fjlkatj': params['z'], 'hy5641d321t': params['O'], 'iogojti': params['S'],
        'ngd4yut78': params['U'], 'nkjhrew': params['h'], 'yt447e13f': params['H'],
        'n3bf4uj7y7': params['B'], 'nbf4uj7y432': params['c'], 'yi854tew': params['c'],
        'h13ey474': params['$'], 'quikgdky': params['C'],
    })
    return data


def _decrypt_with_node(encrypted_text):
    """使用Node.js + crypto-js解密韭圈儿加密数据"""
    try:
        with open(_JQ_TEMP_ENCRYPTED, 'w', encoding='utf-8') as f:
            f.write(encrypted_text)
        env = os.environ.copy()
        env['NODE_PATH'] = _JQ_NODE_PATH
        result = subprocess.run(
            [_JQ_NODE_BIN, _JQ_DECRYPT_SCRIPT, _JQ_TEMP_ENCRYPTED, _JQ_TEMP_DECRYPTED],
            capture_output=True, text=True, env=env, timeout=30
        )
        if result.returncode != 0:
            return None
        with open(_JQ_TEMP_DECRYPTED, 'r', encoding='utf-8') as f:
            data = json.load(f)
        try:
            os.remove(_JQ_TEMP_ENCRYPTED)
            os.remove(_JQ_TEMP_DECRYPTED)
        except:
            pass
        return data
    except Exception:
        return None


def fetch_fear_greed_index():
    """从韭圈儿(jiucaishuo.com) API获取A股恐贪指数
    返回: {"date": "2026-07-24", "value": 8.0, "label": "极度恐惧"} 或 None
    """
    try:
        url = 'https://pre.jiucaishuo.com/v2/kjtl/getbasedata'
        payload = {'formtype': '106000'}
        signed = _sign_jiucaishuo(payload.copy())

        headers = {
            'User-Agent': UA,
            'Content-Type': 'application/x-www-form-urlencoded',
            'Origin': 'https://funddb.cn',
            'Referer': 'https://funddb.cn/tool/fear',
        }
        resp = requests.post(url, data=signed, headers=headers, timeout=15)
        text = resp.text.strip()

        # 如果直接返回JSON(非加密), 说明接口变化了
        if text.startswith('{'):
            try:
                data = json.loads(text)
                if data.get('code') == 0:
                    d = data.get('data', {})
                    fg_value = d.get('num')
                    fg_label = d.get('status_str', '')
                    fg_date = d.get('current_time', '')
                    if fg_value is not None:
                        return {"date": fg_date, "value": int(float(fg_value)), "label": fg_label}
                return None
            except:
                return None

        # AES加密 → Node.js解密
        data = _decrypt_with_node(text)
        if not data or data.get('code') != 0:
            return None

        d = data.get('data', {})
        fg_value = d.get('num')
        fg_label = d.get('status_str', '')
        fg_date = d.get('current_time', '')

        if fg_value is not None:
            return {"date": fg_date, "value": int(float(fg_value)), "label": fg_label}
        return None
    except Exception as e:
        print(f"  [市场状态] 获取恐贪指数失败: {e}")
        return None


def _fetch_mutual_capital_v2(mutual_type, days=5):
    """从东方财富datacenter-web API获取沪深港通资金流向（与基金补仓自动化保持一致）
    mutual_type: '005'北向合计 / '006'南向合计 / '001'沪股通 / '002'深股通
    返回: {"net_flow": float(亿元), "trade_date": "2026-07-24"} 或 None
    """
    url = 'https://datacenter-web.eastmoney.com/api/data/v1/get'
    params = {
        'sortColumns': 'TRADE_DATE',
        'sortTypes': '-1',
        'pageSize': str(days),
        'pageNumber': '1',
        'reportName': 'RPT_MUTUAL_DEAL_HISTORY',
        'columns': 'ALL',
        'filter': f'(MUTUAL_TYPE="{mutual_type}")',
    }
    headers = {
        'User-Agent': UA,
        'Referer': 'https://data.eastmoney.com/',
    }
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=15)
        data = resp.json()
        if data.get('success') and data.get('result'):
            rows = data['result'].get('data', [])
            for row in rows:
                net_amt = row.get('NET_DEAL_AMT')
                if net_amt is not None:
                    trade_date = row.get('TRADE_DATE', '')
                    return {
                        'net_flow': round(net_amt / 1e4, 2),  # 万元 → 亿元
                        'trade_date': trade_date[:10] if trade_date else '',
                    }
    except Exception as e:
        print(f"  [市场状态] 获取沪深港通资金(mutual_type={mutual_type})失败: {e}")
    return None


def fetch_hk_stock_connect():
    """从东方财富datacenter-web获取北向/南向资金（与基金补仓自动化完全一致）
    北向: MUTUAL_TYPE=005（北向合计）；如无数据则用002深股通+001沪股通
    南向: MUTUAL_TYPE=006（南向合计）
    返回: {"north_net": float(亿元), "south_net": float(亿元), "date": str}
    """
    try:
        # 北向资金
        north = _fetch_mutual_capital_v2('005')
        if not north:
            # 回退：深股通(002) + 沪股通(001)
            sz = _fetch_mutual_capital_v2('002')
            sh = _fetch_mutual_capital_v2('001')
            if sz and sh:
                north = {
                    'net_flow': round(sz['net_flow'] + sh['net_flow'], 2),
                    'trade_date': sz['trade_date'] or sh['trade_date'],
                }
            elif sz:
                north = sz

        # 南向资金
        south = _fetch_mutual_capital_v2('006')

        north_net = north['net_flow'] if north else None
        north_date = north['trade_date'] if north else None
        south_net = south['net_flow'] if south else None
        south_date = south['trade_date'] if south else None

        # 日期优先北向，其次南向
        date_str = north_date or south_date

        return {
            "north_net": round(north_net, 2) if north_net is not None else None,
            "south_net": round(south_net, 2) if south_net is not None else None,
            "date": date_str,
        }
    except Exception as e:
        print(f"  [市场状态] 获取沪深港通资金失败: {e}")
        return {"north_net": None, "south_net": None, "date": None}


def fetch_margin_balance():
    """从东方财富datacenter-web获取两市融资余额（与基金补仓自动化保持一致）
    返回: {"date": "2026-07-23", "balance": 26888.61, "net_buy": -97.64,
           "state": "去杠杆", "meaning": "资金收缩", "state_color": "#2E7D32"} 或 None
    balance/net_buy 单位：亿元
    """
    url = 'https://datacenter-web.eastmoney.com/api/data/v1/get'
    params = {
        'sortColumns': 'DIM_DATE',
        'sortTypes': '-1',
        'pageSize': '5',
        'pageNumber': '1',
        'reportName': 'RPTA_RZRQ_LSHJ',
        'columns': 'ALL',
    }
    headers = {
        'User-Agent': UA,
        'Referer': 'https://data.eastmoney.com/',
    }
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=15)
        data = resp.json()
        if data.get('success') and data.get('result'):
            rows = data['result'].get('data', [])
            if rows:
                row = rows[0]
                rzye = row.get('RZYE')
                rzjme = row.get('RZJME')
                dim_date = row.get('DIM_DATE', '')

                balance = round(rzye / 1e8, 2) if rzye is not None else None
                net_buy = round(rzjme / 1e8, 2) if rzjme is not None else None

                # 状态标签：与基金补仓自动化 gen_chart.py 一致
                if net_buy is None:
                    state, meaning, state_color = '—', '—', '#888888'
                elif net_buy > 0:
                    state, meaning, state_color = '加杠杆', '资金进场', '#C62828'
                elif net_buy < 0:
                    state, meaning, state_color = '去杠杆', '资金收缩', '#2E7D32'
                else:
                    state, meaning, state_color = '平稳', '进出均衡', '#7F8C8D'

                return {
                    "date": dim_date[:10] if dim_date else '',
                    "balance": balance,
                    "net_buy": net_buy,
                    "state": state,
                    "meaning": meaning,
                    "state_color": state_color,
                }
    except Exception as e:
        print(f"  [市场状态] 获取融资余额失败: {e}")
    return None


def fetch_sse_index_history(symbol="sh000001", datalen=300):
    """从新浪K线API获取上证综指历史日K数据
    返回: [{"day": "2025-07-22", "open": "...", "close": "...", "high": "...", "low": "..."}, ...]
    """
    try:
        url = f"https://money.finance.sina.com.cn/quotes_service/api/json_v2.php/CN_MarketData.getKLineData?symbol={symbol}&scale=240&datalen={datalen}"
        r = requests.get(url, headers={'User-Agent': UA}, timeout=20)
        data = r.json()
        if isinstance(data, list) and len(data) > 0:
            return data
        print(f"  [市场状态] 上证综指K线数据异常: {str(data)[:200]}")
        return []
    except Exception as e:
        print(f"  [市场状态] 获取上证综指历史数据失败: {e}")
        return []


def get_confirmed_market_turnover(target_date):
    """从同日盘中快照读取沪深全市场成交额，避免把日K成交量误当成交额。"""
    snapshot_path = os.path.join(os.path.dirname(__file__), "deploy", "live_estimates.json")
    try:
        with open(snapshot_path, "r", encoding="utf-8") as f:
            snapshot = json.load(f)
        if snapshot.get("trade_date") != target_date:
            return None
        value = (snapshot.get("sse_index") or {}).get("market_turnover_trillion")
        return float(value) if value is not None else None
    except (OSError, ValueError, json.JSONDecodeError):
        return None


def detect_market_regime(danjuan_data):
    """检测当前A股市场状态（4维度: 均线趋势/估值分位/RSI/回撤深度）
    
    返回: (regime_key, regime_info)
      regime_key: "early_bull" | "mid_bull" | "late_bull" | "correction"
    """
    reasons = []
    
    # --- 获取上证综指300日K线 ---
    klines = fetch_sse_index_history("sh000001", 300)
    if not klines or len(klines) < 200:
        print("  [市场状态] K线数据不足（需>=200日），使用默认中牛市参数")
        return "mid_bull", {
            "name": "牛市中期(默认)", "desc": "K线数据不足，使用默认参数",
            "classification": "数据不足", "reasons": ["K线数据<200日"],
            "sse_close": None, "sse_ma200": None, "sse_above_ma200": None,
            "sse_rsi": None, "sse_drawdown_from_peak": None,
            "sse_ma200_pct": None, "hs300_pe_pct": None,
            "sse_prev_close": None, "sse_change_pct": None, "sse_change_amount": None,
            "market_turnover_trillion": None,
        }
    
    closes = [float(k["close"]) for k in klines]
    latest_close = closes[-1]
    prev_close = closes[-2] if len(closes) >= 2 else latest_close
    sse_change_amount = latest_close - prev_close
    sse_change_pct = (latest_close - prev_close) / prev_close * 100 if prev_close else 0
    # 新浪日K的 volume 字段是成交量，不是成交额；确认层仅使用同日盘中快照保存的沪深全市场成交额。
    market_turnover_trillion = get_confirmed_market_turnover(klines[-1].get("day"))
    
    # 维度1: 200日均线
    ma200 = sum(closes[-200:]) / 200
    above_ma200 = latest_close > ma200
    ma200_pct = (latest_close - ma200) / ma200 * 100
    
    # 维度3: RSI(14)
    sse_rsi = calc_rsi_from_closes(closes, 14)
    if sse_rsi is None:
        sse_rsi = 50  # 默认中性
    
    # 维度4: 从近期峰值回撤
    recent = closes[-250:] if len(closes) >= 250 else closes
    peak = max(recent)
    peak_idx = recent.index(peak)
    drawdown_from_peak = (latest_close - peak) / peak * 100  # 负数
    days_since_peak = len(recent) - peak_idx - 1
    
    # 维度2: 沪深300 PE分位
    hs300_pe_pct = None
    dj_code = DANJUAN_CODE_MAP.get("000300")  # "SH000300"
    if dj_code and dj_code in danjuan_data:
        pct_raw = danjuan_data[dj_code].get("pe_percentile")
        if pct_raw is not None:
            hs300_pe_pct = pct_raw * 100  # 蛋卷返回的是小数，转百分比
    
    # --- 分类逻辑 ---
    if not above_ma200:
        if drawdown_from_peak < -20:
            regime_key = "correction"
            reasons.append(f"跌破200日线({ma200_pct:+.1f}%)且峰值回撤{drawdown_from_peak:.1f}%，深度调整")
        else:
            regime_key = "correction"
            reasons.append(f"跌破200日线({ma200_pct:+.1f}%)，回撤{drawdown_from_peak:.1f}%，牛市调整中")
    elif drawdown_from_peak > -5 and sse_rsi > 60:
        if hs300_pe_pct and hs300_pe_pct > 80:
            regime_key = "late_bull"
            reasons.append(f"在200日线上方({ma200_pct:+.1f}%)，RSI={sse_rsi}偏高，PE分位={hs300_pe_pct:.0f}%偏高→牛市后期")
        else:
            regime_key = "mid_bull"
            reasons.append(f"在200日线上方({ma200_pct:+.1f}%)，RSI={sse_rsi}，估值适中")
    elif drawdown_from_peak > -15:
        regime_key = "mid_bull"
        reasons.append(f"在200日线上方({ma200_pct:+.1f}%)，峰值回撤{drawdown_from_peak:.1f}%，牛市中期")
    else:
        if sse_rsi < 40:
            regime_key = "early_bull"
            reasons.append(f"在200日线上方({ma200_pct:+.1f}%)，RSI={sse_rsi}低位，可能处于熊转牛初期")
        else:
            regime_key = "mid_bull"
            reasons.append(f"在200日线上方({ma200_pct:+.1f}%)，RSI={sse_rsi}")
    
    # --- 打印诊断信息 ---
    regime_params = MARKET_REGIME_TABLE[regime_key]
    print(f"\n  [市场状态] {'='*50}")
    print(f"  [市场状态] 上证综指={latest_close:.2f} | 200日线={ma200:.2f} | 偏离{ma200_pct:+.1f}%")
    print(f"  [市场状态] RSI(14)={sse_rsi} | 峰值回撤={drawdown_from_peak:.1f}% (距峰值{days_since_peak}天)")
    if hs300_pe_pct is not None:
        print(f"  [市场状态] 沪深300 PE分位={hs300_pe_pct:.1f}%")
    print(f"  [市场状态] → {regime_params['name']}: {regime_params['desc']}")
    
    for r in reasons:
        print(f"  [市场状态]   原因: {r}")
    
    regime_info = {
        "regime_key": regime_key,
        "name": regime_params["name"],
        "desc": regime_params["desc"],
        "sse_close": round(latest_close, 2),
        "sse_ma200": round(ma200, 2),
        "sse_above_ma200": above_ma200,
        "sse_ma200_pct": round(ma200_pct, 1),
        "sse_rsi": sse_rsi,
        "sse_drawdown_from_peak": round(drawdown_from_peak, 1),
        "sse_days_since_peak": days_since_peak,
        "sse_prev_close": round(prev_close, 2),
        "sse_change_pct": round(sse_change_pct, 2),
        "sse_change_amount": round(sse_change_amount, 2),
        "market_turnover_trillion": round(market_turnover_trillion, 2) if market_turnover_trillion is not None else None,
        "hs300_pe_pct": round(hs300_pe_pct, 1) if hs300_pe_pct else None,
        "reasons": reasons,
    }

    # --- 获取额外市场情绪指标 ---
    fg = fetch_fear_greed_index()
    hk = fetch_hk_stock_connect()
    margin = fetch_margin_balance()

    regime_info["fear_greed"] = fg
    regime_info["hk_connect"] = hk
    regime_info["margin_balance"] = margin

    if fg:
        print(f"  [市场状态] 恐贪指数={fg['value']} ({fg['label']}) 日期={fg['date']}")
    if hk.get("north_net") is not None:
        print(f"  [市场状态] 北向={hk['north_net']:+.2f}亿 南向={hk['south_net']:+.2f}亿 日期={hk['date']}")
    if margin:
        balance_trillion = margin['balance'] / 1e4
        print(f"  [市场状态] 融资余额={balance_trillion:.2f}万亿元 净买入={margin['net_buy']:+.2f}亿 ({margin['state']}) 日期={margin['date']}")

    return regime_key, regime_info


def get_regime_tp_params(regime_key):
    """根据市场状态返回自适应止盈参数"""
    regime = MARKET_REGIME_TABLE.get(regime_key, MARKET_REGIME_TABLE["mid_bull"])
    return {
        "rsi_tp_strong": regime["rsi_tp_strong"],
        "rsi_tp_val": regime["rsi_tp_val"],
        "rsi_tp_x_gain": regime["rsi_tp_x_gain"],
        "sell_pct_scale": regime["sell_pct_scale"],
        "anchor_clear_rsi": regime["anchor_clear_rsi"],
    }


def get_index_valuation(index_code, danjuan_data):
    """
    综合获取指数估值数据（蛋卷优先 → 腾讯备选 → 中证官网PE → legulegu PB）
    返回 {"pe": float, "pb": float, "source": str, "pb_flag": bool, "pe_pct": float, "pb_pct": float}
    """
    result = {"pe": 0, "pb": 0, "source": "N/A", "pb_flag": False, "pe_pct": None, "pb_pct": None}

    if index_code is None:
        return result

    # 1. 尝试蛋卷API
    dj_code = DANJUAN_CODE_MAP.get(index_code)
    if dj_code and dj_code in danjuan_data:
        d = danjuan_data[dj_code]
        result["pe"] = d["pe"] if d["pe"] is not None else 0
        result["pb"] = d["pb"] if d["pb"] is not None else 0
        result["pb_flag"] = d.get("pb_flag", False)
        result["pe_pct"] = d.get("pe_percentile")
        result["pb_pct"] = d.get("pb_percentile")
        result["source"] = "蛋卷基金"
        return result

    # 2. 尝试腾讯财经
    tencent = get_index_pe_pb_tencent(index_code)
    if tencent and (tencent["pe"] > 0 or tencent["pb"] > 0):
        result["pe"] = tencent["pe"]
        result["pb"] = tencent["pb"]
        result["source"] = "腾讯财经"
        return result

    # 3. 尝试中证官网perf API PE（针对类型二指数）
    pe_perf = fetch_csindex_pe_perf(index_code)
    if pe_perf and pe_perf.get("pe", 0) > 0:
        result["pe"] = pe_perf["pe"]
        result["source"] = "中证官网"
        # 同时尝试PB
        pb_lg = fetch_legulegu_pb(index_code)
        if pb_lg and pb_lg.get("pb", 0) > 0:
            result["pb"] = pb_lg["pb"]
            result["source"] = "中证(PE)+乐咕(PB)"
        return result

    # 4. 尝试legulegu PB（单独获取）
    pb_lg = fetch_legulegu_pb(index_code)
    if pb_lg and pb_lg.get("pb", 0) > 0:
        result["pb"] = pb_lg["pb"]
        result["source"] = "乐咕乐股"
        return result

    # 5. 无法获取
    result["source"] = None
    return result


def classify_with_vref(index_code, pe, pb, use_pb=False):
    """
    根据VALUATION_REF阈值分类估值水平
    use_pb: True → 强制用PB分类, False → 用PE分类（如果PE有效）
    返回 (signal, metric_name, metric_value)
    """
    ref = VALUATION_REF.get(index_code)
    if not ref:
        # 无参考阈值，用通用阈值
        if use_pb and pb > 0:
            if pb < 0.8: return (f"PB={pb:.2f} 跌破净值", "PB", pb)
            elif pb < 1.5: return (f"PB={pb:.2f} 低估", "PB", pb)
            elif pb < 3.0: return (f"PB={pb:.2f} 适中", "PB", pb)
            elif pb < 5.0: return (f"PB={pb:.2f} 偏高", "PB", pb)
            else: return (f"PB={pb:.2f} 高估", "PB", pb)
        elif pe > 0 and pe < 500:
            if pe < 15: return (f"PE={pe:.1f} 低估", "PE", pe)
            elif pe < 25: return (f"PE={pe:.1f} 适中", "PE", pe)
            elif pe < 50: return (f"PE={pe:.1f} 偏高", "PE", pe)
            else: return (f"PE={pe:.1f} 高估", "PE", pe)
        return ("数据不足", "N/A", None)

    # 1. 强制使用PB：ref只有"pb"键
    if "pb" in ref and "pe" not in ref:
        lo, mid, hi = ref["pb"]
        desc = ref.get("desc", "")
        if pb <= 0:
            # PB不可用，检查PE是否可作备选
            if pe > 0 and pe < 500:
                if pe < 15: return (f"PE={pe:.1f} 偏低[PB首选不可用]", "PE", pe)
                elif pe < 25: return (f"PE={pe:.1f} 适中[PB首选不可用]", "PE", pe)
                elif pe < 50: return (f"PE={pe:.1f} 偏高[PB首选不可用]", "PE", pe)
                else: return (f"PE={pe:.1f} 高估[PB首选不可用]", "PE", pe)
            na_reason = f"需PB估值但PB数据缺失({desc})"
            return (na_reason, "PB", None)
        if pb <= lo: return (f"PB={pb:.2f} 低估", "PB", pb)
        elif pb <= mid: return (f"PB={pb:.2f} 适中", "PB", pb)
        elif pb <= hi: return (f"PB={pb:.2f} 偏高", "PB", pb)
        else: return (f"PB={pb:.2f} 高估", "PB", pb)

    # 2. 强制使用PB：use_pb=True (pb_flag或PE>200)
    if use_pb and "pb" in ref:
        lo, mid, hi = ref["pb"]
        if pb <= 0:
            # PB不可用，检查PE备选
            if pe > 0 and pe < 500 and "pe" in ref:
                pe_lo, pe_mid, pe_hi = ref["pe"]
                if pe <= pe_lo: return (f"PE={pe:.1f} 偏低[PB首选不可用]", "PE", pe)
                elif pe <= pe_mid: return (f"PE={pe:.1f} 适中[PB首选不可用]", "PE", pe)
                elif pe <= pe_hi: return (f"PE={pe:.1f} 偏高[PB首选不可用]", "PE", pe)
                else: return (f"PE={pe:.1f} 高估[PB首选不可用]", "PE", pe)
            return ("PB数据缺失", "PB", None)
        if pb <= lo: return (f"PB={pb:.2f} 低估", "PB", pb)
        elif pb <= mid: return (f"PB={pb:.2f} 适中", "PB", pb)
        elif pb <= hi: return (f"PB={pb:.2f} 偏高", "PB", pb)
        else: return (f"PB={pb:.2f} 高估", "PB", pb)

    # 2.5. use_pb=True 但ref无pb键 → 用通用PB阈值
    if use_pb:
        if pb > 0:
            if pb < 0.8: return ("跌破净值", "PB", pb)
            elif pb < 1.5: return ("低估", "PB", pb)
            elif pb < 3.0: return ("适中", "PB", pb)
            elif pb < 5.0: return ("偏高", "PB", pb)
            else: return ("高估", "PB", pb)
        else:
            return ("PB数据缺失", "PB", None)

    # 3. 使用PE分类
    lo, mid, hi = ref["pe"]
    if pe <= 0:
        if use_pb and pb > 0:
            # PE无效但有PB和ref.pb
            return ("PE亏损[改用PB]", "PB", pb)
        return ("PE为负/无效", "PE", None)
    if pe <= lo: return (f"PE={pe:.1f} 低估", "PE", pe)
    elif pe <= mid: return (f"PE={pe:.1f} 适中", "PE", pe)
    elif pe <= hi: return (f"PE={pe:.1f} 偏高", "PE", pe)
    else: return (f"PE={pe:.1f} 高估", "PE", pe)


def get_valuation_info_v2(code, category, index_code, index_val):
    """
    V2版本的估值信息 (K列)
    返回 (val_metric, val_value, val_signal_desc, val_na_reason, val_detail)
    """
    # --- 商品类基金 ---
    if "黄金" in category:
        return ("不适用", None, "商品基金无PE/PB估值",
                "黄金为商品类资产，不适用市盈率/市净率估值", None)

    # --- 主动管理型基金 ---
    if index_code is None:
        return ("不适用", None, "主动基金无指数估值",
                f"主动管理型基金（{category}），无跟踪指数，PE/PB来自基金持仓而非指数", None)

    # --- 获取指数估值 ---
    pe = index_val["pe"]
    pb = index_val["pb"]
    pb_flag = index_val["pb_flag"]
    source = index_val["source"]

    # --- 确定应该用PE还是PB ---
    ref = VALUATION_REF.get(index_code, {})

    # 规则1: ref只有"pb"键 → 必须用PB
    pb_only_ref = "pb" in ref and "pe" not in ref

    # 规则2: 蛋卷pb_flag=True → 建议用PB
    # 规则3: PE ≤ 0 → PE无效
    # 规则4: PE > 200 → PE异常，改用PB

    use_pb = pb_only_ref or pb_flag or (pe <= 0) or (pe > 200)

    if source is None:
        # 无法获取估值数据
        reasons = []
        if code == "024194":
            reasons.append("国证卫星通信指数(980018)不在中证/蛋卷/腾讯/乐咕任意免费数据源")
            reasons.append("中证官网不覆盖国证指数，国证官网API不可用")
        elif code == "018344":
            # H30590应在perf API获取到PE，如果到这里说明网络异常
            reasons.append("机器人指数(H30590)所有数据源均获取失败")
        elif code == "013402":
            pass  # 恒生科技在蛋卷里
        else:
            reasons.append(f"指数({index_code})不在任何估值数据源(蛋卷/腾讯/中证/乐咕)")

        reason = "; ".join(reasons)

        # 特殊说明：如果应该用PB但PB获取不到
        if pb_only_ref:
            return ("不适用", None, f"PB数据源不可用",
                    f"需用PB估值({reason})，建议查看理杏仁/中证指数官网", reason)
        return ("不适用", None, f"指数PE/PB数据暂缺",
                reason, reason)

    # --- 分类估值 ---
    signal, metric_name, metric_val = classify_with_vref(index_code, pe, pb, use_pb=use_pb)

    # --- 构建详细说明 ---
    detail_parts = [f"数据源: {source}"]
    if pb_only_ref:
        if pe <= 0:
            detail_parts.append("行业亏损PE失效，强制用PB")
        else:
            detail_parts.append(f"周期行业应看PB(当前PE={pe:.1f}可用但不准确)")
    elif pb_flag:
        detail_parts.append("蛋卷标记pb_flag=True，建议用PB")
    elif pe > 200:
        detail_parts.append(f"PE={pe:.1f}异常(>200)，改用PB")
    elif pe <= 0:
        detail_parts.append(f"PE≤0无效")
    else:
        # 正常PE估值，但检查是否为PB首选不可用的情况
        if pb <= 0 and "pb" in ref and "pe" in ref:
            detail_parts.append(f"PB首选但数据不可用，PE作为参考(指标={metric_name})")
        else:
            detail_parts.append("正常PE估值")

    if index_val["pe_pct"] is not None:
        detail_parts.append(f"PE分位={index_val['pe_pct']:.1%}")
    if index_val["pb_pct"] is not None:
        detail_parts.append(f"PB分位={index_val['pb_pct']:.1%}")

    detail = " | ".join(detail_parts)
    return (metric_name, metric_val, signal, None, detail)


# ===== 信号判断函数 =====
def get_level(drawdown_pct):
    for name, lo, hi, mul in LEVEL_TABLE:
        if lo <= drawdown_pct < hi:
            return name, mul
    return "极限3级", 7


def get_rsi_signal(rsi):
    if rsi is None: return "N/A"
    if rsi < 20: return "极度超卖"
    if rsi < 30: return "超卖"
    if rsi < 45: return "偏弱"
    if rsi <= 70: return "中性"
    return "超买"


# 综合信号/操作建议颜色配置（M列字体颜色 + L列可选背景）
SIGNAL_COLORS = {
    "强烈补仓": {"text": "1B5E20", "fill": "C6EFCE"},   # 深绿
    "建议补仓": {"text": "2E7D32", "fill": "C6EFCE"},   # 绿色
    "可补仓":   {"text": "E65100", "fill": "FFEB9C"},   # 橙色
    "关注":     {"text": "5D6D7E", "fill": "D9E2EC"},   # 蓝灰
    "观望":     {"text": "7F8C8D", "fill": "D9D9D9"},   # 灰色
    "暂不操作": {"text": "95A5A6", "fill": "D9D9D9"},   # 浅灰
    "警惕回调": {"text": "C62828", "fill": "FFC7CE"},   # 红色
    "考虑止盈": {"text": "8B0000", "fill": "FFE0B2"},   # 暗红（区别于警惕回调的亮红）
}


def classify_valuation_for_signal(val_signal):
    """把估值描述简化为三档，用于综合信号判断"""
    if val_signal is None:
        return "na"
    v = str(val_signal)
    if "偏低" in v or "低估" in v or "跌破净值" in v or "适中" in v:
        return "good"
    if "偏高" in v or "高估" in v:
        return "bad"
    return "na"


def get_signal_and_advice(drawdown_pct, rsi, val_signal, trend_20d_pct=None, regime_params=None):
    """
    三指标综合信号判断矩阵（V2.2：趋势判断+止盈信号）
    V2.3新增：市场状态自适应止盈阈值（regime_params）

    返回三元组：
      - comp_meaning: K列"综合信号"含义文字
      - action_text:  L列"操作建议"文字（无emoji前缀）
      - action_color: M列字体颜色配置 dict {text, fill}

    趋势判断：
      - trend_20d_pct > 0 → 反弹中，买入信号自动降一级
      - trend_20d_pct <= 0 → 下跌中，信号保持不变

    止盈信号（最高优先级，不被趋势降级影响）：
      - RSI > regime_params['rsi_tp_strong'] → "考虑止盈"（默认75，牛市初期80，后期65）
      - RSI > regime_params['rsi_tp_val'] + 估值偏高/高估 → "考虑止盈"
      调整期(correction)时阈值设为999，实际不触发止盈
    """
    val_cls = classify_valuation_for_signal(val_signal)
    is_recovering = trend_20d_pct is not None and trend_20d_pct > 0

    # ========== 止盈信号（最高优先级，先于所有买入信号） ==========
    # V2.3: 使用市场状态自适应阈值
    tp_strong = regime_params['rsi_tp_strong'] if regime_params else 75
    tp_val_thresh = regime_params['rsi_tp_val'] if regime_params else 70

    if rsi is not None:
        rsi_tp_strong = rsi > tp_strong
        rsi_tp_val = rsi > tp_val_thresh and val_cls == "bad"
        if rsi_tp_strong:
            return f"RSI超高位({rsi}>{tp_strong})，考虑分批止盈", "考虑止盈", SIGNAL_COLORS["考虑止盈"]
        if rsi_tp_val:
            return f"估值偏高+RSI超买(>{tp_val_thresh})，考虑止盈", "考虑止盈", SIGNAL_COLORS["考虑止盈"]

    rsi_oversold = rsi is not None and rsi < 30
    rsi_neutral = rsi is not None and 30 <= rsi <= 70
    rsi_overbought = rsi is not None and rsi > tp_val_thresh  # 70-阈值且估值不bad，未被止盈捕获

    # <13% 停止区
    if drawdown_pct < 13:
        return "趋势完好，无需补仓", "暂不操作", SIGNAL_COLORS["暂不操作"]

    # 13-15% 观望区
    if drawdown_pct < 15:
        return "接近警戒线，可关注等待", "暂不操作", SIGNAL_COLORS["暂不操作"]

    # 15-20% 倍投1级 — 估值不影响判断
    if drawdown_pct < 20:
        if rsi_overbought:
            return "价格已反弹，不宜追高", "警惕回调", SIGNAL_COLORS["警惕回调"]
        if rsi_oversold:
            meaning = "回撤尚浅，等跌幅加深"
            action, color = "观望", SIGNAL_COLORS["观望"]
            if is_recovering:
                meaning += "（反弹中）"
            return meaning, action, color
        # 30-70 中性
        meaning = "信号不充分，持续观察"
        action, color = "观望", SIGNAL_COLORS["观望"]
        if is_recovering:
            meaning += "（反弹中）"
        return meaning, action, color

    # 20-33% 倍投2-3级 — 估值不影响判断
    if drawdown_pct < 33:
        if rsi_overbought:
            return "价格已反弹，不宜追高", "警惕回调", SIGNAL_COLORS["警惕回调"]
        if rsi_oversold:
            meaning = "回撤达标+RSI超卖，择机补仓"
            action, color = "可补仓", SIGNAL_COLORS["可补仓"]
            if is_recovering:
                meaning += "（反弹中降级）"
                action, color = "关注", SIGNAL_COLORS["关注"]
            return meaning, action, color
        # 30-70 中性
        meaning = "回撤达标但RSI未超卖，等待"
        action, color = "关注", SIGNAL_COLORS["关注"]
        if is_recovering:
            meaning += "（反弹中）"
        return meaning, action, color

    # 33%+ 倍投4级 至 极限3级 — 估值开始区分
    # ===== Zone A: 33-55% (倍投4级 ~ 极限1级) — 原有逻辑不变 =====
    if drawdown_pct < 55:
        if rsi_overbought:
            return "深跌后已反弹，不宜追高", "警惕回调", SIGNAL_COLORS["警惕回调"]
        if rsi_oversold:
            if val_cls == "bad":
                meaning = "回撤深+RSI超卖，但估值偏高需谨慎"
                action, color = "可补仓", SIGNAL_COLORS["可补仓"]
                if is_recovering:
                    meaning += "（反弹中降级）"
                    action, color = "关注", SIGNAL_COLORS["关注"]
                return meaning, action, color
            if val_cls == "good":
                meaning = "三指标共振，强烈建议补仓"
                action, color = "强烈补仓", SIGNAL_COLORS["强烈补仓"]
                if is_recovering:
                    meaning += "（反弹中降级）"
                    action, color = "建议补仓", SIGNAL_COLORS["建议补仓"]
                return meaning, action, color
            # 不适用 / 黄金 / 主动基金 / 卫星
            meaning = "回撤深+RSI超卖（黄金/主动基金适用）"
            action, color = "建议补仓", SIGNAL_COLORS["建议补仓"]
            if is_recovering:
                meaning += "（反弹中降级）"
                action, color = "可补仓", SIGNAL_COLORS["可补仓"]
            return meaning, action, color
        # 30-70 中性
        if val_cls == "bad":
            meaning = "回撤深但估值高+RSI中性，等待确认"
            action, color = "关注", SIGNAL_COLORS["关注"]
            return meaning, action, color
        # 低估/适中/不适用
        meaning = "回撤深但RSI未超卖，择机补仓"
        action, color = "可补仓", SIGNAL_COLORS["可补仓"]
        if is_recovering:
            meaning += "（反弹中降级）"
            action, color = "关注", SIGNAL_COLORS["关注"]
        return meaning, action, color

    # ===== Zone B: 55%+ (极限2-3级) — 信号增强 =====
    # 深度回撤conviction提升：所有信号比33-55%区系统性+1级
    if rsi_overbought:
        return "深跌后已大幅反弹，不宜追高", "警惕回调", SIGNAL_COLORS["警惕回调"]
    if rsi_oversold:
        if val_cls == "bad":
            meaning = "回撤极深+RSI超卖，但估值偏高需谨慎"
            action, color = "建议补仓", SIGNAL_COLORS["建议补仓"]  # ↑从可补仓提升
            if is_recovering:
                meaning += "（反弹中降级）"
                action, color = "可补仓", SIGNAL_COLORS["可补仓"]
            return meaning, action, color
        if val_cls == "good":
            meaning = "三指标共振，历史级底部机会"
            action, color = "强烈补仓", SIGNAL_COLORS["强烈补仓"]  # 同级，文字增强
            if is_recovering:
                meaning += "（反弹中降级）"
                action, color = "建议补仓", SIGNAL_COLORS["建议补仓"]
            return meaning, action, color
        # 不适用 / 黄金 / 主动基金 / 卫星
        meaning = "回撤极深+RSI超卖，历史性底部区域"
        action, color = "强烈补仓", SIGNAL_COLORS["强烈补仓"]  # ↑从建议补仓提升
        if is_recovering:
            meaning += "（反弹中降级）"
            action, color = "建议补仓", SIGNAL_COLORS["建议补仓"]
        return meaning, action, color
    # 30-70 中性
    if val_cls == "bad":
        meaning = "回撤极深但估值偏高，谨慎关注"
        action, color = "可补仓", SIGNAL_COLORS["可补仓"]  # ↑从关注提升
        if is_recovering:
            meaning += "（反弹中降级）"
            action, color = "关注", SIGNAL_COLORS["关注"]
        return meaning, action, color
    # 低估/适中/不适用
    meaning = "回撤极深+估值低位，择机补仓"
    action, color = "建议补仓", SIGNAL_COLORS["建议补仓"]  # ↑从可补仓提升
    if is_recovering:
        meaning += "（反弹中降级）"
        action, color = "可补仓", SIGNAL_COLORS["可补仓"]
    return meaning, action, color


# ===== 止盈锚点持久化 =====
def load_tp_state():
    """加载止盈锚点状态"""
    try:
        with open(TP_STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_tp_state(state):
    """保存止盈锚点状态"""
    with open(TP_STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def read_trough_data():
    """从Excel模板读取V/W列底谷数据，返回 {code: (trough_date, trough_nav)}
    作为v6 JSON中trough数据的回退源（向后兼容）"""
    try:
        wb = load_workbook(EXCEL_TEMPLATE_PATH, data_only=True)
        ws = wb['金字塔丛林补仓']
        result = {}
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row=row, column=2).value
            if not code:
                continue
            code = str(code).strip()
            v_val = ws.cell(row=row, column=22).value   # V列底谷日期
            w_val = ws.cell(row=row, column=23).value   # W列底谷净值
            if w_val is not None:
                try:
                    w_float = float(w_val)
                    if w_float > 0:
                        result[code] = (str(v_val) if v_val else None, w_float)
                except (ValueError, TypeError):
                    pass
        wb.close()
        return result
    except Exception as e:
        print(f"  [WARN] 读取底谷数据失败: {e}")
        return {}


def get_take_profit_info(rsi, val_signal_desc, current_nav, x_gain_pct, code, tp_state, regime_params=None):
    """
    止盈策略（三维度触发 + X列等级修正 + 市场状态自适应）

    V2.3: 触发阈值和卖出比例根据市场状态自适应调整
      - early_bull: 阈值提高(80/75)，卖出比例×0.6
      - mid_bull: 阈值正常(75/70)，卖出比例×1.0
      - late_bull: 阈值降低(65/60)，卖出比例×1.5
      - correction: 阈值999(不触发)，卖出比例×0

    触发条件:
      ① RSI > rsi_tp_strong → 无条件触发
      ② RSI > rsi_tp_val + 估值偏高/高估 → 估值风险触发
      ③ RSI > rsi_tp_x_gain + X列涨幅 > 25% → 短期涨幅风险触发

    等级计算:
      基础等级 = 涨幅vs止盈锚点
      X列修正: <15%降1级 / 15-30%不变 / >30%升1级
      sell_pct_scale: 最终卖出比例 × 缩放系数

    锚点机制:
      首次触发 → 设锚点=当前净值，无建议（观察中）
      已有锚点 → 计算涨幅，返回止盈建议
      信号消失(RSI < anchor_clear_rsi) → 清除锚点

    返回: (tp_anchor, tp_anchor_date, tp_level, tp_gain_pct, tp_sell_pct, tp_display)
    """
    if rsi is None:
        return None, None, None, None, None, None

    # V2.3: 使用市场状态自适应参数
    if regime_params:
        tp_strong = regime_params['rsi_tp_strong']
        tp_val_thresh = regime_params['rsi_tp_val']
        tp_x_gain_thresh = regime_params['rsi_tp_x_gain']
        sell_pct_scale = regime_params['sell_pct_scale']
        anchor_clear_rsi = regime_params['anchor_clear_rsi']
    else:
        tp_strong = 75
        tp_val_thresh = 70
        tp_x_gain_thresh = 70
        sell_pct_scale = 1.0
        anchor_clear_rsi = 50

    val_cls = classify_valuation_for_signal(val_signal_desc)

    # 检查是否触发止盈（V2.3: 使用自适应阈值）
    trigger = (
        rsi > tp_strong or
        (rsi > tp_val_thresh and val_cls == "bad") or
        (rsi > tp_x_gain_thresh and x_gain_pct is not None and x_gain_pct > 25)
    )

    fund_state = tp_state.get(code, {})

    if not trigger:
        # 信号未触发：检查是否需要清除锚点（V2.3: 自适应清除阈值）
        if fund_state.get("anchor") is not None:
            if rsi < anchor_clear_rsi:
                del tp_state[code]
                print(f"  [止盈] {code} RSI回落至{rsi}(<{anchor_clear_rsi})，清除止盈锚点")
        return None, None, None, None, None, None

    # === 止盈信号已触发 ===

    # 情况1：首次触发，设置锚点
    if fund_state.get("anchor") is None:
        fund_state["anchor"] = current_nav
        fund_state["anchor_date"] = datetime.now().strftime("%Y-%m-%d")
        tp_state[code] = fund_state
        print(f"  [止盈] {code} 首次触发止盈信号，锚点={current_nav:.4f}")
        # 观察区不返回建议
        return current_nav, fund_state["anchor_date"], None, 0.0, 0, None

    # 情况2：已有锚点，计算涨幅和止盈等级
    anchor = fund_state["anchor"]
    anchor_date = fund_state.get("anchor_date", "?")
    gain_pct = round((current_nav - anchor) / anchor * 100, 2)

    # 查找基础止盈等级
    base_level_name = None
    base_sell_pct = 0
    for name, lo, hi, sell in TAKE_PROFIT_LEVELS:
        if lo <= abs(gain_pct) < hi or (lo <= abs(gain_pct) and hi == 999):
            base_level_name = name
            base_sell_pct = sell
            break

    if base_level_name is None:
        base_level_name = "清仓区"
        base_sell_pct = 100

    # 观察区不显示
    if base_level_name == "观察区":
        return anchor, anchor_date, None, gain_pct, 0, None

    # X列涨幅修正止盈等级
    level_names = [l[0] for l in TAKE_PROFIT_LEVELS]
    level_idx = level_names.index(base_level_name)

    if x_gain_pct is not None:
        if x_gain_pct < 15:
            # 刚涨不久，保守卖 → 降1级
            level_idx = max(level_idx - 1, 1)  # 不低于止盈1级
        elif x_gain_pct > 30:
            # 涨多了，积极卖 → 升1级
            level_idx = min(level_idx + 1, len(level_names) - 1)  # 不高于清仓区

    final_level_name = level_names[level_idx]
    final_sell_pct = TAKE_PROFIT_LEVELS[level_idx][3]

    # V2.3: 应用市场状态缩放系数（调整卖出比例）
    if sell_pct_scale != 1.0:
        final_sell_pct = int(round(final_sell_pct * sell_pct_scale))
        final_sell_pct = min(final_sell_pct, 100)  # 不超过100%
        # 观察区(0%)缩放后仍为0
        if final_level_name == "观察区":
            final_sell_pct = 0

    # 生成显示文字（含市场状态标识）
    regime_tag = ""
    if regime_params and sell_pct_scale != 1.0:
        regime_tag = f"(×{sell_pct_scale})"

    tp_display = f"{final_level_name}·卖{final_sell_pct}%{regime_tag}"

    if final_level_name == "清仓区":
        tp_display = f"{final_level_name}·卖{final_sell_pct}%{regime_tag}"

    return anchor, anchor_date, final_level_name, gain_pct, final_sell_pct, tp_display


def fetch_latest_nav_and_change(code):
    url = 'http://api.fund.eastmoney.com/f10/lsjz'
    params = {'callback': '', 'fundCode': code, 'pageIndex': '1', 'pageSize': '3',
              'startDate': '', 'endDate': ''}
    try:
        r = requests.get(url, params=params, headers=API_HEADERS, timeout=15)
        d = r.json()
        items = d.get('Data', {}).get('LSJZList', [])
        if not items:
            return None, None, None, None
        latest = items[0]
        nav = float(latest['DWJZ'])
        date = latest['FSRQ']
        change_pct = float(latest.get('JZZZL', '0') or '0')
        prev_nav = float(items[1]['DWJZ']) if len(items) > 1 else None
        return nav, date, change_pct, prev_nav
    except Exception as e:
        print(f"  [{code}] lsjz ERROR: {e}")
        return None, None, None, None


def fetch_recent_nav_history(code, count=20):
    """获取最近N个交易日的确认净值序列，用于RSI实时重算
    返回: [nav_t-1, nav_t-2, ...] 从新到旧（仅含已确认净值，不含当天）
    """
    url = 'http://api.fund.eastmoney.com/f10/lsjz'
    all_navs = []
    page = 1
    while len(all_navs) < count and page <= 10:
        page_size = min(20, count - len(all_navs))
        params = {'callback': '', 'fundCode': code, 'pageIndex': str(page),
                  'pageSize': str(page_size), 'startDate': '', 'endDate': ''}
        try:
            r = requests.get(url, params=params, headers=API_HEADERS, timeout=15)
            d = r.json()
            items = d.get('Data', {}).get('LSJZList', [])
            if not items:
                break
            for item in items:
                try:
                    all_navs.append(float(item['DWJZ']))
                except (KeyError, ValueError):
                    pass
            if len(items) < page_size:
                break
            page += 1
        except Exception as e:
            print(f"  [{code}] 拉历史净值失败(page={page}): {e}")
            break
    return all_navs[:count]


def calc_rsi_with_today(historical_navs, estimated_nav, period=14):
    """用历史确认净值 + 当天估算净值重新计算RSI(14)
    historical_navs: [T-1净值, T-2净值, ...] lsjz已确认，从新到旧
    estimated_nav: 当天估算净值（Sina fu_接口）
    返回: RSI值 或 None
    """
    # 序列: [estimated_nav(today), T-1, T-2, ..., T-period]
    all_navs = [estimated_nav] + historical_navs
    if len(all_navs) < period + 1:
        return None

    # 取最近period+1个值（含当天）
    navs = all_navs[:period + 1]
    # 计算逐日涨跌: navs[i-1] - navs[i]（navs[0]=今天, navs[1]=昨天）
    changes = [navs[i] - navs[i + 1] for i in range(len(navs) - 1)]
    changes = changes[:period]  # 只取period个变化

    gains = sum(c for c in changes if c > 0) / period
    losses = sum(-c for c in changes if c < 0) / period
    if losses == 0:
        return 100.0
    rs = gains / losses
    return round(100 - 100 / (1 + rs), 1)


def calc_rsi_series(navs, period=14):
    """从净值序列计算滚动RSI序列（用于折线图）
    navs: list of NAV values, oldest→newest
    Returns: list of RSI values, same order as input (oldest→newest).
    前period个位置为None（数据不足以计算RSI）。
    """
    if len(navs) < period + 1:
        return [None] * len(navs)
    
    result = [None] * len(navs)
    for i in range(period, len(navs)):
        window = navs[i - period:i + 1]
        changes = [window[j] - window[j - 1] for j in range(1, len(window))]
        gains = sum(c for c in changes if c > 0) / period
        losses = sum(-c for c in changes if c < 0) / period
        if losses == 0:
            result[i] = 100.0
        else:
            result[i] = round(100 - 100 / (1 + gains / losses), 1)
    return result


def fetch_sina_fund_estimate(code):
    """从新浪财经获取基金实时估值涨跌幅（fu_接口，覆盖全部场外基金）
    返回格式: var hq_str_fu_000218="国泰黄金ETF联接A,16:04:00,3.2333,3.2024,3.2024,0,0.9649,2026-07-22,3.2362,1.0555";
    字段: [0]名称 [1]估值时间 [2]估算净值 [3]最新确认净值 [4]累计净值 [5]0 [6]估算涨跌幅% [7]日期 [8]前一日净值 [9]前一日涨跌幅%
    返回: (change_pct, source_info) 或 (None, None)
    """
    try:
        url = f"https://hq.sinajs.cn/list=fu_{code}"
        r = requests.get(url, headers={
            'User-Agent': UA, 'Referer': 'https://finance.sina.com.cn/'
        }, timeout=10)
        text = r.content.decode('gbk', errors='replace')

        if f'hq_str_fu_{code}' not in text or '=' not in text:
            return None, None

        data_str = text.split('"')[1]
        if not data_str or len(data_str) < 10:
            return None, None

        fields = data_str.split(',')
        name = fields[0]
        est_nav = float(fields[2])
        confirmed_nav = float(fields[3])
        date = fields[7]
        est_time = fields[1]

        if confirmed_nav == 0 or est_nav == 0:
            return None, None

        change_pct = (est_nav - confirmed_nav) / confirmed_nav * 100
        source = f"Sina估值({name},{date} {est_time})"
        return round(change_pct, 2), source
    except Exception as e:
        print(f"  Sina fu_{code} ERROR: {e}")
        return None, None


def fetch_sina_realtime_change(sina_code):
    """从新浪财经获取指数/ETF的实时涨跌幅
    返回: (change_pct, source_info) 或 (None, None)
    """
    try:
        url = f"https://hq.sinajs.cn/list={sina_code}"
        r = requests.get(url, headers={
            'User-Agent': UA, 'Referer': 'https://finance.sina.com.cn/'
        }, timeout=10)
        # Sina返回GBK编码
        text = r.content.decode('gbk', errors='replace')

        if '=' not in text or len(text) < 50:
            return None, None

        # 解析: var hq_str_XXX="name,open,prev_close,current,high,low,..."
        data_str = text.split('"')[1]
        if not data_str or len(data_str) < 10:
            return None, None

        fields = data_str.split(',')
        name = fields[0]

        # field[2]=昨收(prev_close), field[3]=当前价(current)
        current = float(fields[3])
        prev_close = float(fields[2])

        if prev_close == 0 or current == 0:
            return None, None

        change_pct = (current - prev_close) / prev_close * 100
        source = f"Sina实时({name})"
        return round(change_pct, 2), source
    except Exception as e:
        print(f"  Sina {sina_code} ERROR: {e}")
        return None, None


def get_realtime_daily_change(code, index_code):
    """获取基金当天涨跌幅
    优先级: Sina基金估值(fu_,覆盖全部) → Sina指数实时 → Sina ETF实时 → None(回退到lsjz)
    返回: (change_pct, source) 或 (None, None)
    """
    # 1. Sina基金估值接口（fu_覆盖全部场外基金，包括黄金/主动/电网等）
    change, source = fetch_sina_fund_estimate(code)
    if change is not None:
        return change, source

    # 2. Sina指数实时行情
    if index_code and index_code in SINA_CODE_MAP:
        sina_code = SINA_CODE_MAP[index_code]
        change, source = fetch_sina_realtime_change(sina_code)
        if change is not None:
            return change, source

    # 3. Sina ETF实时行情（CSI自定义指数）
    if index_code and index_code in ETF_CODE_MAP:
        etf_code = ETF_CODE_MAP[index_code]
        change, source = fetch_sina_realtime_change(etf_code)
        if change is not None:
            return change, source

    # 4. 无法获取，回退到lsjz
    return None, None


def check_trading_day():
    """判断今天是否为A股交易日
    
    逻辑：
    1. 周末(Sat/Sun) → 非交易日
    2. 节假日：通过Sina fu_ API检测 → 如数据日期≠今天 → 非交易日
    
    返回: (is_trading, note)
    """
    today_str = date.today().strftime('%Y-%m-%d')
    today_weekday = date.today().weekday()
    weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    
    # 周末必非交易日
    if today_weekday >= 5:
        return False, f"周末休市({weekday_names[today_weekday]})"
    
    # 工作日：检查Sina是否返回当天数据
    try:
        r = requests.get("https://hq.sinajs.cn/list=fu_005453", headers={
            'User-Agent': UA, 'Referer': 'https://finance.sina.com.cn/'
        }, timeout=10)
        text = r.content.decode('gbk', errors='replace')
        if 'hq_str_fu_005453' in text and '"' in text:
            data_str = text.split('"')[1]
            fields = data_str.split(',')
            if len(fields) >= 8:
                api_date = fields[7]
                if api_date != today_str:
                    return False, f"节假日休市(Sina日期={api_date})"
        return True, "交易日"
    except Exception as e:
        return True, f"交易日(检测异常: {e})"


def previous_weekday(today):
    """返回最近一个工作日。法定节假日由后续逐基金净值日期校验兜底。"""
    candidate = today - timedelta(days=1)
    while candidate.weekday() >= 5:
        candidate -= timedelta(days=1)
    return candidate


def parse_args():
    parser = argparse.ArgumentParser(description='金字塔丛林基金数据增强')
    parser.add_argument(
        '--final-nav',
        action='store_true',
        help='确认净值模式：仅接受指定目标日的确认净值；F列日内估算涨跌将清空。'
    )
    parser.add_argument(
        '--target-nav-date',
        help='确认净值目标日，格式YYYY-MM-DD；未指定时默认取最近一个工作日。'
    )
    return parser.parse_args()


def main():
    args = parse_args()
    final_nav_mode = args.final_nav
    today = date.today()
    expected_nav_date = previous_weekday(today)
    run_date = args.target_nav_date or expected_nav_date.isoformat()
    mode_name = '早间确认净值模式' if final_nav_mode else '盘中实时估值模式'
    print(f'运行模式: {mode_name} | 目标净值日期: {run_date}')

    # ----- Step 0: 检测交易日状态 -----
    print("=" * 70)
    print("Step 0: 检测交易日状态...")
    is_trading_day, trading_note = check_trading_day()
    print(f"  今日状态: {trading_note}")
    print()

    # ----- Step 1: 获取蛋卷指数估值数据 -----
    print("=" * 70)
    print("Step 1: 获取蛋卷指数估值数据...")
    danjuan_data = fetch_danjuan_index_valuation()
    print(f"  蛋卷API返回 {len(danjuan_data)} 个指数的估值数据")
    matched_count = sum(1 for c in DANJUAN_CODE_MAP.values() if c in danjuan_data)
    print(f"  其中 {matched_count} 个匹配我们的跟踪指数")

    # ----- Step 1b: 检测市场状态（Regime Detection） -----
    print("\nStep 1b: 检测A股市场状态...")
    regime_key, regime_info = detect_market_regime(danjuan_data)
    regime_params = get_regime_tp_params(regime_key)
    print(f"  → 自适应止盈参数: 触发阈值={regime_params['rsi_tp_strong']}/{regime_params['rsi_tp_val']}/{regime_params['rsi_tp_x_gain']}, 卖出系数={regime_params['sell_pct_scale']}, 清锚阈值={regime_params['anchor_clear_rsi']}")

    # ----- Step 2: 读取V6基础数据 -----
    print("\nStep 2: 读取V6基础数据...")
    with open('fund_data_v6.json', 'r', encoding='utf-8') as f:
        v6_data = json.load(f)

    # 读取底谷数据（V/W列）和止盈锚点状态
    print("Step 2b: 读取Excel底谷数据(V/W列)...")
    trough_data = read_trough_data()
    print(f"  底谷数据: {len(trough_data)} 只基金")
    tp_state = load_tp_state()
    print(f"  止盈锚点状态: {len(tp_state)} 只基金有锚点")

    funds = [(code, info) for code, info in v6_data.items() if not info.get('error')]
    funds.sort(key=lambda x: x[1].get('drawdown_pct', 0), reverse=True)

    enriched = []
    unconfirmed_codes = []
    total = len(funds)

    print(f"\nStep 3: 开始处理 {total} 只基金...")
    print("=" * 70)

    for idx, (code, old) in enumerate(funds):
        name = old.get('name', code)
        category = old.get('category', '')
        index_code = INDEX_MAP.get(code)
        print(f"\n[{idx+1}/{total}] {code} {name} [{category}]")

        # 获取指数估值数据
        index_val = get_index_valuation(index_code, danjuan_data)
        print(f"  指数={index_code} | 数据源={index_val['source']} | PE={index_val['pe']} PB={index_val['pb']} pb_flag={index_val['pb_flag']}")

        # 获取最新净值和日涨跌
        latest_nav, latest_date, lsjz_change, prev_nav = fetch_latest_nav_and_change(code)
        time.sleep(0.3)

        # 早间最终模式只接受目标日的确认净值；不依赖当前时点的盘中交易日判断。
        # 早上9点尚未开盘，实时接口通常仍显示前一交易日，不能误判为“非交易日”。
        nav_confirmed_today = bool(latest_date == run_date)
        if final_nav_mode and not nav_confirmed_today:
            reason = f'确认净值日期={latest_date or "缺失"}，尚未等于目标日{run_date}'
            print(f"  ⏳ 未确认目标日净值，跳过写入：{reason}")
            unconfirmed_codes.append(code)
            continue

        # F列仅供盘中估值展示。早间确认净值模式清空，避免与E列终值重复计入G/Q/X。
        if final_nav_mode:
            daily_change = None
            change_source = f"确认净值({latest_date})"
        elif is_trading_day:
            realtime_change, realtime_source = get_realtime_daily_change(code, index_code)
            if realtime_change is not None:
                daily_change = realtime_change
                change_source = realtime_source
            else:
                daily_change = lsjz_change
                change_source = f"lsjz结算({latest_date})"
        else:
            daily_change = None
            change_source = trading_note

        if latest_nav is None:
            print(f"  ❌ 无法获取净值数据，跳过")
            continue

        # 峰值和回撤
        all_time_high = old.get('all_time_high', latest_nav)
        high_date = old.get('high_date', '')
        drawdown_pct = round((all_time_high - latest_nav) / all_time_high * 100, 1)

        # 底谷数据（优先用V6计算值，回退到Excel V/W列）
        trough_date = old.get('trough_date')
        trough_nav = old.get('trough_nav')
        if trough_nav is None and code in trough_data:
            trough_date, trough_nav = trough_data[code]
            print(f"  底谷(Excel回退): {trough_nav:.4f} ({trough_date})" if trough_nav else "  底谷: 无数据")
        elif trough_nav is not None:
            print(f"  底谷(V6计算): {trough_nav:.4f} ({trough_date})")

        # 倍投等级
        level_name, multiplier = get_level(drawdown_pct)

        # RSI - 用当天估算净值重算（含当日涨跌幅）
        rsi_old = old.get('rsi')
        rsi = rsi_old  # 默认用V6的T-1 RSI
        rsi_source = "T-1(lsjz)"

        # 拉取45日确认净值历史（RSI用前14日+当天，趋势判断用20日，折线图用足够历史）
        recent_navs = fetch_recent_nav_history(code, count=45)
        time.sleep(0.2)

        # RSI历史序列（用于折线图）—— 从确认净值计算。
        # recent_navs 按 newest→oldest 排列；当前确认净值 RSI 必须与折线末值同源。
        rsi_history = None
        # 保留完整45个确认净值点：正式简单RSI仍只取最近15点，
        # 但独立Wilder验证需要更长的平滑暖机历史。该字段不参与Excel公式或止盈锚点写入。
        confirmed_navs_desc = recent_navs[:45] if recent_navs else []
        if recent_navs:
            navs_old_to_new = list(reversed(recent_navs))
            rsi_history_raw = calc_rsi_series(navs_old_to_new, 14)
            rsi_history = [v for v in rsi_history_raw if v is not None]
            confirmed_rsi = calc_rsi_with_today(recent_navs[1:], recent_navs[0], 14)
            if confirmed_rsi is not None:
                rsi = confirmed_rsi
                rsi_source = f"确认净值({latest_date})"

        if not final_nav_mode and 'Sina' in change_source and daily_change is not None and recent_navs:
            # 当天估算净值 = 最新确认净值 × (1 + 当天涨跌幅%)
            estimated_nav = round(latest_nav * (1 + daily_change / 100), 4)
            rsi_new = calc_rsi_with_today(recent_navs, estimated_nav, 14)
            if rsi_new is not None:
                rsi = rsi_new
                rsi_source = f"含当日估算({daily_change:+.2f}%)"
                print(f"  RSI: {rsi} (含当日{daily_change:+.2f}%, 确认RSI={confirmed_rsi})")
                # 在确认净值 RSI 序列末尾追加含当日估算的最新 RSI 值。
                if rsi_history is not None:
                    rsi_history.append(rsi_new)
            else:
                print(f"  RSI: {rsi} (重算失败, 回退确认RSI)")
        else:
            print(f"  RSI: {rsi} ({rsi_source})")
        rsi_signal = get_rsi_signal(rsi)

        # 趋势判断：严格取 21 个确认净值点，即最近 20 个交易日变化。
        trend_20d_pct = None
        trend_status = "未知"
        trend_navs = recent_navs[:21] if recent_navs else []
        if len(trend_navs) >= 2:
            nav_latest = trend_navs[0]
            nav_oldest = trend_navs[-1]
            if nav_oldest > 0:
                trend_20d_pct = round((nav_latest - nav_oldest) / nav_oldest * 100, 2)
                if trend_20d_pct > 0:
                    trend_status = "反弹中"
                elif trend_20d_pct < 0:
                    trend_status = "下跌中"
                else:
                    trend_status = "横盘"
        trend_info = f"{trend_status}({trend_20d_pct:+.2f}%)" if trend_20d_pct is not None else "数据不足"
        print(f"  趋势: {trend_info}")

        # J列: PE/PB估值 (V2 智能选择)
        val_metric, val_value, val_signal_desc, val_na_reason, val_detail = \
            get_valuation_info_v2(code, category, index_code, index_val)

        # 打印估值详情
        if daily_change is not None:
            print(f"  涨跌={daily_change:+.2f}% ({change_source})")
        else:
            print(f"  涨跌=休市 ({change_source})")
        print(f"  估值: {val_metric}={val_value} → {val_signal_desc}")

        # 综合信号(K列含义) + 操作建议(L列信号文字+颜色)
        comp_meaning, action_text, action_color = get_signal_and_advice(
            drawdown_pct, rsi, val_signal_desc, trend_20d_pct, regime_params=regime_params)

        # X列涨幅计算 — 使用已设置的trough_nav（V6优先，Excel回退）
        # 注意：不要从trough_data重新读取，否则会用Excel旧数据覆盖V6新数据
        x_gain_pct = None
        if trough_nav and trough_nav > 0:
            x_gain_pct = round((latest_nav - trough_nav) / trough_nav * 100, 2)

        tp_anchor, tp_anchor_date, tp_level, tp_gain_pct, tp_sell_pct, tp_display = \
            get_take_profit_info(rsi, val_signal_desc, latest_nav, x_gain_pct, code, tp_state, regime_params=regime_params)
        if tp_display:
            print(f"  [止盈] {tp_display} (涨幅vs锚点={tp_gain_pct:+.1f}%, X列={x_gain_pct}%)")

        # 建议补仓份数
        floor_dd = int(drawdown_pct)
        suggested_shares = floor_dd * multiplier * BASE

        entry = {
            'code': code,
            'name': name,
            'category': category,
            'index_code': index_code,
            'latest_nav': latest_nav,
            'latest_date': latest_date,
            'final_nav_mode': final_nav_mode,
            'nav_confirmed_today': nav_confirmed_today,
            'daily_change': None if daily_change is None else round(daily_change, 2),
            'daily_change_source': change_source,
            'all_time_high': all_time_high,
            'high_date': high_date,
            'drawdown_pct': drawdown_pct,
            'level': level_name,
            'multiplier': multiplier,
            'rsi': rsi,
            'rsi_signal': rsi_signal,
            'rsi_source': rsi_source,
            'rsi_as_of_date': latest_date,
            'rsi_history_end_date': latest_date,
            'confirmed_navs_desc': confirmed_navs_desc,  # newest→oldest，供盘中RSI(14)重算
            'rsi_history': rsi_history,  # RSI历史序列 (oldest→newest)，用于折线图
            # K列
            'val_metric': val_metric,
            'val_value': val_value,
            'val_signal': val_signal_desc,
            'val_na_reason': val_na_reason,
            'val_detail': val_detail,
            'val_source': index_val['source'],
            'val_pe': index_val['pe'] if index_val['pe'] > 0 else None,
            'val_pb': index_val['pb'] if index_val['pb'] > 0 else None,
            'val_pb_flag': index_val['pb_flag'],
            'val_pe_pct': index_val['pe_pct'],
            'val_pb_pct': index_val['pb_pct'],
            # K列: 综合信号（含义文字）
            'comp_meaning': comp_meaning,
            # L列: 操作建议（信号文字 + 颜色）
            'comp_signal': action_text,
            'comp_signal_color': action_color,
            # 趋势判断
            'trend_20d_pct': trend_20d_pct,
            'trend_status': trend_status,
            # 止盈策略（Y列）
            'tp_anchor': tp_anchor,
            'tp_anchor_date': tp_anchor_date,
            'tp_level': tp_level,
            'tp_gain_pct': tp_gain_pct,
            'tp_sell_pct': tp_sell_pct,
            'tp_display': tp_display,
            'x_gain_pct': x_gain_pct,
            'trough_date': trough_date,
            'trough_nav': trough_nav,
            'suggested_shares': suggested_shares,
            'level_str': f"{multiplier}x" if multiplier == int(multiplier) else f"{multiplier}x",
        }
        enriched.append(entry)

    # 早间发布必须全量确认，不能用部分基金的旧快照覆盖正式JSON/Excel/公开页。
    if final_nav_mode and unconfirmed_codes:
        print(f"\n⚠️ 早间确认净值不完整：{len(unconfirmed_codes)}/{total} 只未确认（{', '.join(unconfirmed_codes)}）")
        print("已停止保存与发布；请在下一轮重试，避免把旧净值当作当日最终净值。")
        raise SystemExit(2)

    # ----- Step 4: 保存 -----
    output = {
        'update_time': time.strftime('%Y-%m-%d %H:%M:%S'),
        'data_mode': 'final_nav' if final_nav_mode else 'intraday_estimate',
        'target_nav_date': run_date,
        'fund_count': len(enriched),
        'total_fund_count': total,
        'unconfirmed_codes': unconfirmed_codes,
        'market_regime': regime_info,
        'market_regime_params': regime_params,
        'funds': enriched,
    }

    with open('fund_data_enriched.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # 保存止盈锚点状态
    save_tp_state(tp_state)
    tp_active = sum(1 for v in tp_state.values() if v.get('anchor') is not None)
    print(f"  止盈锚点状态已保存: {tp_active} 只基金有活跃锚点")

    print("\n" + "=" * 70)
    print(f"✅ 完成！{len(enriched)}/{total} 只基金数据已保存到 fund_data_enriched.json")

    # ----- 汇总统计 -----
    print("\n📊 PE/PB估值分布:")
    pe_count = sum(1 for e in enriched if e['val_metric'] == 'PE')
    pb_count = sum(1 for e in enriched if e['val_metric'] == 'PB')
    na_count = sum(1 for e in enriched if e['val_metric'] == '不适用')
    print(f"  PE估值: {pe_count}只 | PB估值: {pb_count}只 | 不适用: {na_count}只")

    print("\n🔍 估值信号分布:")
    signals = {}
    for e in enriched:
        s = f"{e['val_metric']}:{e['val_signal']}"
        signals[s] = signals.get(s, 0) + 1
    for s, c in sorted(signals.items(), key=lambda x: x[1], reverse=True):
        print(f"  {s}: {c}只")

    print("\n📊 等级分布:")
    for lv_name, lo, hi, mul in LEVEL_TABLE:
        cnt = sum(1 for e in enriched if e['level'] == lv_name)
        bar = '█' * cnt
        print(f"  {lv_name} ({lo}-{hi}%): {bar} {cnt}只")

    print("\n" + "=" * 70)
    print("各基金K列估值详情:")
    print("-" * 70)
    for e in enriched:
        na_info = f" | {e['val_na_reason']}" if e['val_na_reason'] else ""
        print(f"  {e['code']} {e['name']}: {e['val_metric']}={e['val_value']} → {e['val_signal']}{na_info}")


if __name__ == '__main__':
    main()
