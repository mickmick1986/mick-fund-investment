#!/usr/bin/env python3
"""Generate a non-invasive Simple RSI vs Wilder RSI validation dataset.

This module reads the confirmed fund JSON and Excel inputs but never writes the
formal Excel workbook, fund_data_enriched.json, or take_profit_state.json.
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent
INPUT_PATH = ROOT / "fund_data_enriched.json"
OUTPUT_PATH = ROOT / "rsi_dual_track_validation.json"
PUBLIC_OUTPUT_PATH = ROOT / "deploy" / "rsi_dual_track_validation.json"
HISTORY_PATH = ROOT / "rsi_dual_track_history.json"
EXCEL_PATH = Path(r"C:\Users\13697\Desktop\金字塔丛林战法\基金模板（金字塔丛林版）.xlsx")
SHEET_NAME = "金字塔丛林补仓"
BUY_SIGNALS = {"强烈补仓", "建议补仓", "可补仓"}


def rsi_signal(rsi: float | None) -> str:
    if rsi is None:
        return "N/A"
    if rsi < 20:
        return "极度超卖"
    if rsi < 30:
        return "超卖"
    if rsi < 45:
        return "偏弱"
    if rsi <= 70:
        return "中性"
    return "超买"


def calc_simple_rsi(closes: list[float], period: int = 14) -> float | None:
    if len(closes) < period + 1:
        return None
    changes = [closes[i] - closes[i - 1] for i in range(1, len(closes))]
    changes = changes[-period:]
    gains = sum(change for change in changes if change > 0) / period
    losses = sum(-change for change in changes if change < 0) / period
    if losses == 0:
        return 100.0
    return round(100 - 100 / (1 + gains / losses), 1)


def calc_wilder_rsi_series(closes: list[float], period: int = 14) -> list[float | None]:
    """Classic Wilder RSI series, oldest to newest.

    The initial average is seeded from the first 14 changes, then each following
    price uses Wilder's recursive smoothing formula.
    """
    result: list[float | None] = [None] * len(closes)
    if len(closes) < period + 1:
        return result
    changes = [closes[i] - closes[i - 1] for i in range(1, len(closes))]
    avg_gain = sum(change for change in changes[:period] if change > 0) / period
    avg_loss = sum(-change for change in changes[:period] if change < 0) / period

    def to_rsi(gain: float, loss: float) -> float:
        if loss == 0:
            return 100.0
        return round(100 - 100 / (1 + gain / loss), 1)

    result[period] = to_rsi(avg_gain, avg_loss)
    for index in range(period + 1, len(closes)):
        change = changes[index - 1]
        gain = max(change, 0)
        loss = max(-change, 0)
        avg_gain = (avg_gain * (period - 1) + gain) / period
        avg_loss = (avg_loss * (period - 1) + loss) / period
        result[index] = to_rsi(avg_gain, avg_loss)
    return result


def get_level(drawdown_pct: float) -> tuple[str, float]:
    levels = [
        ("停止区", 0, 13, 0), ("观望区", 13, 15, 0.5), ("倍投1级", 15, 20, 1),
        ("倍投2级", 20, 25, 1.5), ("倍投3级", 25, 33, 2), ("倍投4级", 33, 40, 3),
        ("倍投5级", 40, 48, 4), ("极限1级", 48, 55, 5), ("极限2级", 55, 65, 6),
        ("极限3级", 65, 999, 7),
    ]
    for name, low, high, multiplier in levels:
        if low <= drawdown_pct < high:
            return name, multiplier
    return "极限3级", 7


def valuation_class(value: str | None) -> str:
    text = str(value or "")
    if any(item in text for item in ("偏低", "低估", "跌破净值", "适中")):
        return "good"
    if any(item in text for item in ("偏高", "高估")):
        return "bad"
    return "na"


def simulated_advice(drawdown: float, rsi: float | None, valuation: str, trend: float | None, params: dict[str, Any]) -> str:
    """Decision-equivalent mirror of the formal three-indicator action only."""
    val_cls = valuation_class(valuation)
    recovering = trend is not None and trend > 0
    strong = params.get("rsi_tp_strong", 75)
    value_threshold = params.get("rsi_tp_val", 70)
    oversold = rsi is not None and rsi < 30
    # 验证轨镜像正式规则：买入侧超买固定为RSI>70，不复用止盈阈值。
    overbought = rsi is not None and rsi > 70
    if rsi is not None and (rsi > strong or (rsi > value_threshold and val_cls == "bad")):
        return "考虑止盈"
    if drawdown < 15:
        return "暂不操作"
    if drawdown < 20:
        return "观望"
    if drawdown < 33:
        if overbought:
            return "警惕回调"
        return "关注" if not oversold or recovering else "可补仓"
    if drawdown < 55:
        if overbought:
            return "警惕回调"
        if oversold:
            if val_cls == "bad":
                return "关注" if recovering else "可补仓"
            if val_cls == "good":
                return "建议补仓" if recovering else "强烈补仓"
            return "可补仓" if recovering else "建议补仓"
        return "关注" if val_cls == "bad" or recovering else "可补仓"
    if overbought:
        return "警惕回调"
    if oversold:
        if val_cls == "bad":
            return "可补仓" if recovering else "建议补仓"
        if val_cls == "good":
            return "建议补仓" if recovering else "强烈补仓"
        return "建议补仓" if recovering else "强烈补仓"
    return "关注" if val_cls == "bad" and recovering else ("可补仓" if recovering or val_cls == "bad" else "建议补仓")


def projected_order(action: str, multiplier: float, anchor_nav: float | None, nav: float, unit_price: float | None) -> dict[str, float | int]:
    if action not in BUY_SIGNALS:
        return {"shares": 0, "multiple": 0, "amount": 0}
    anchor_change = (nav - anchor_nav) / anchor_nav if anchor_nav else 0
    shares = max(int(-anchor_change * 100), 1)
    multiple = round(multiplier * shares, 1)
    amount = round((unit_price or 0) * multiple, 2)
    return {"shares": shares, "multiple": multiple, "amount": amount}


def load_excel_inputs() -> dict[str, dict[str, float | None]]:
    if not EXCEL_PATH.exists():
        return {}
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet = workbook[SHEET_NAME]
    result: dict[str, dict[str, float | None]] = {}
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 2).value
        if value is None:
            continue
        code = str(value).zfill(6)
        result[code] = {
            "anchor_nav": sheet.cell(row, 15).value,
            "unit_price": sheet.cell(row, 19).value,
        }
    workbook.close()
    return result


def trigger_take_profit(rsi: float | None, valuation: str, x_gain: float | None, params: dict[str, Any]) -> bool:
    if rsi is None:
        return False
    return bool(
        rsi > params.get("rsi_tp_strong", 75)
        or (rsi > params.get("rsi_tp_val", 70) and valuation_class(valuation) == "bad")
        or (rsi > params.get("rsi_tp_x_gain", 70) and x_gain is not None and x_gain > 25)
    )


def main() -> int:
    data = json.loads(INPUT_PATH.read_text(encoding="utf-8"))
    excel = load_excel_inputs()
    params = data.get("market_regime_params", {})
    records = []
    for fund in data.get("funds", []):
        code = str(fund["code"]).zfill(6)
        closes_desc = [float(value) for value in fund.get("confirmed_navs_desc", []) if value is not None]
        closes = list(reversed(closes_desc))
        wilder_series = calc_wilder_rsi_series(closes)
        wilder = next((value for value in reversed(wilder_series) if value is not None), None)
        simple = fund.get("rsi")
        simple = round(float(simple), 1) if simple is not None else calc_simple_rsi(closes)
        drawdown = float(fund.get("drawdown_pct", 0))
        level, multiplier = get_level(drawdown)
        formal_action = fund.get("comp_signal", "")
        wilder_action = simulated_advice(drawdown, wilder, fund.get("val_signal", ""), fund.get("trend_20d_pct"), params)
        excel_input = excel.get(code, {})
        simple_order = projected_order(formal_action, multiplier, excel_input.get("anchor_nav"), float(fund["latest_nav"]), excel_input.get("unit_price"))
        wilder_order = projected_order(wilder_action, multiplier, excel_input.get("anchor_nav"), float(fund["latest_nav"]), excel_input.get("unit_price"))
        formal_tp = trigger_take_profit(simple, fund.get("val_signal", ""), fund.get("x_gain_pct"), params)
        wilder_tp = trigger_take_profit(wilder, fund.get("val_signal", ""), fund.get("x_gain_pct"), params)
        usable_points = sum(value is not None for value in wilder_series)
        records.append({
            "code": code,
            "name": fund.get("name", code),
            "category": fund.get("category", ""),
            "as_of_date": fund.get("rsi_as_of_date") or fund.get("latest_date"),
            "simple_rsi": simple,
            "wilder_rsi": wilder,
            "rsi_delta": round(wilder - simple, 1) if wilder is not None and simple is not None else None,
            "simple_signal": rsi_signal(simple),
            "wilder_signal": rsi_signal(wilder),
            "formal_action": formal_action,
            "wilder_action": wilder_action,
            "level": level,
            "simple_order": simple_order,
            "wilder_order": wilder_order,
            "formal_tp_trigger": formal_tp,
            "wilder_tp_trigger": wilder_tp,
            "formal_tp_display": fund.get("tp_display"),
            "decision_changed": formal_action != wilder_action,
            "buy_changed": simple_order["shares"] != wilder_order["shares"],
            "tp_trigger_changed": formal_tp != wilder_tp,
            "wilder_series": [value for value in wilder_series if value is not None],
            "simple_series": fund.get("rsi_history", []),
            "history_point_count": len(closes),
            "wilder_usable_points": usable_points,
        })

    records.sort(key=lambda item: (not item["decision_changed"], item["code"]))
    result = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_update_time": data.get("update_time"),
        "source_mode": data.get("data_mode"),
        "as_of_date": data.get("target_nav_date"),
        "validation_mode": "confirmed_only",
        "formal_track": "现行简单平均RSI(14)，唯一正式决策来源",
        "candidate_track": "Wilder平滑RSI(14)，仅用于验证和模拟",
        "integrity_note": "本文件不会回写Excel、确认层JSON或止盈锚点。",
        "warmup_note": "Wilder使用可获得的确认净值序列；45点历史可提供更稳定的暖机，当前旧快照不足45点时页面会显示实际点数。",
        "summary": {
            "fund_count": len(records),
            "action_changed_count": sum(item["decision_changed"] for item in records),
            "buy_changed_count": sum(item["buy_changed"] for item in records),
            "tp_trigger_changed_count": sum(item["tp_trigger_changed"] for item in records),
            "same_action_count": sum(not item["decision_changed"] for item in records),
        },
        "funds": records,
    }
    serialized_result = json.dumps(result, ensure_ascii=False, indent=2)
    OUTPUT_PATH.write_text(serialized_result, encoding="utf-8")
    PUBLIC_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_OUTPUT_PATH.write_text(serialized_result, encoding="utf-8")

    history = []
    if HISTORY_PATH.exists():
        try:
            history = json.loads(HISTORY_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            history = []
    snapshot_key = f"{result['as_of_date']}|{result['source_mode']}"
    history = [entry for entry in history if entry.get("snapshot_key") != snapshot_key]
    history.append({"snapshot_key": snapshot_key, "generated_at": result["generated_at"], "summary": result["summary"], "funds": records})
    HISTORY_PATH.write_text(json.dumps(history[-120:], ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Generated validation for {len(records)} funds: action differences={result['summary']['action_changed_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
