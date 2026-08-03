#!/usr/bin/env python3
"""
生成金字塔丛林补仓指导图 HTML
数据源：Excel(最终数值) + enriched JSON(信号/颜色/RSI)
列分区：
  - 基金信息区：A(类型), B(代码), C(名称)
  - 三指标判断详情区：G(总回撤), H(RSI), J(估值), L(操作建议), M(倍投等级)
  - 当日补仓指引区：F(当日涨跌), Q(近期总涨跌幅), R(补仓份数), S(补仓单价), T(补仓倍数), U(补仓金额)
  - 近期止盈参考区：V(底谷日期), W(底谷净值), X(上涨幅度), Y(止盈建议)
"""
import json
import os
import openpyxl
from datetime import date

EXCEL_PATH = r'C:\Users\13697\Desktop\金字塔丛林战法\基金模板（金字塔丛林版）.xlsx'
JSON_PATH = 'fund_data_enriched.json'

# CI模式检测：如果设置了CI_MODE环境变量，使用repo内相对路径
if os.environ.get('CI_MODE', '').lower() == 'true':
    EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '基金模板（金字塔丛林版）.xlsx')
SHEET_NAME = '金字塔丛林补仓'

def load_json():
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    rows = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 2).value
        if not code:
            continue
        rows.append({
            'code': str(code).zfill(6),
            'cat': ws.cell(r, 1).value or '',
            'name': ws.cell(r, 3).value or '',
            # 公式源数据（非展示列，用于compute_formulas）:
            'd_val': ws.cell(r, 4).value,            # D: 历史最高净值
            'e_val': ws.cell(r, 5).value,            # E: 最新净值
            'o_val': ws.cell(r, 15).value,           # O: 锚点净值
            # 脚本写入/手动填写列:
            'rsi': ws.cell(r, 8).value,             # H: RSI
            'rsi_signal': ws.cell(r, 9).value,       # I: RSI信号
            'val_display': ws.cell(r, 10).value,     # J: 估值
            'comp_meaning': ws.cell(r, 11).value,    # K: 综合信号
            'action': ws.cell(r, 12).value,          # L: 操作建议
            'level': ws.cell(r, 13).value,           # M: 等级
            'grade_mult': ws.cell(r, 14).value,      # N: 倍投基数
            'daily_change': ws.cell(r, 6).value,     # F: 当日涨跌
            'unit_price': ws.cell(r, 19).value,      # S: 补仓单价
            'trough_date': ws.cell(r, 22).value,     # V: 底谷日期
            'trough_nav': ws.cell(r, 23).value,      # W: 底谷净值
            'tp_signal': ws.cell(r, 25).value,       # Y: 止盈建议
            'group': ws.cell(r, 29).value or '',     # AC: 分组
            'fundamental': ws.cell(r, 30).value or '', # AD: 基本面状态
        })
    return rows

def compute_formulas(excel_rows):
    """在Python中计算静态确认净值公式列（data_only=True拿不到缓存值）。
    F列是盘中展示字段，不能参与静态G/Q/X计算，否则E已为最终净值时会重复计入。"""
    for er in excel_rows:
        e = er.get('e_val')  # E: 最新净值
        d = er.get('d_val')  # D: 历史最高净值
        f_val = er.get('daily_change')  # F: 当日涨跌
        o_val = er.get('o_val')  # O: 锚点净值
        w_val = er.get('trough_nav')  # W: 底谷净值
        n_val = er.get('grade_mult')  # N: 倍投基数
        s_val = er.get('unit_price')  # S: 补仓单价
        action = er.get('action', '')  # L: 操作建议
        
        def to_float(v):
            if v is None:
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None
        
        e_n = to_float(e)
        d_n = to_float(d)
        f_n = to_float(f_val)
        o_n = to_float(o_val)
        w_n = to_float(w_val)
        n_n = to_float(n_val)
        s_n = to_float(s_val)
        
        # G = (E-D)/D；E是确认净值，静态计算不叠加F。
        if e_n is not None and d_n is not None and d_n != 0:
            g_val = (e_n - d_n) / d_n
        else:
            g_val = None
        er['drawdown'] = g_val
        
        # P = (E-O)/O
        if e_n is not None and o_n is not None and o_n != 0:
            p_val = (e_n - o_n) / o_n
        else:
            p_val = None
        er['p_val'] = p_val
        
        # Q = P = (E-O)/O；锚点比较直接以确认净值计算。
        q_val = p_val
        er['q_val'] = q_val
        
        # R = IF(OR(L="强烈补仓",L="建议补仓",L="可补仓"),MAX(IF(Q<0,INT(-Q*100),0),1),0)
        buy_signals = ['强烈补仓', '建议补仓', '可补仓']
        if action in buy_signals and q_val is not None and q_val < 0:
            r_val = max(int(-q_val * 100), 1)
        elif action in buy_signals:
            r_val = 1  # 信号有效时保底1份
        else:
            r_val = 0
        er['shares'] = r_val
        
        # T = N * R
        if n_n is not None and r_val is not None:
            t_val = n_n * r_val
        else:
            t_val = None
        er['final_mult'] = t_val
        
        # U = S * T
        if s_n is not None and t_val is not None:
            u_val = s_n * t_val
        else:
            u_val = None
        er['amount'] = u_val
        
        # X = (E-W)/W；底谷涨幅直接以确认净值计算。
        if e_n is not None and w_n is not None and w_n != 0:
            x_val = (e_n - w_n) / w_n
        else:
            x_val = None
        er['x_gain'] = x_val


def merge_data(excel_rows, json_data):
    """合并Excel数值和JSON颜色/信号"""
    json_map = {f['code']: f for f in json_data['funds']}
    merged = []
    for er in excel_rows:
        jf = json_map.get(er['code'], {})
        m = dict(er)
        # Excel L列为权威来源，颜色从action文字本身推导，不用JSON覆盖
        m['rsi_signal_j'] = jf.get('rsi_signal', er.get('rsi_signal', ''))
        m['rsi_history'] = jf.get('rsi_history', [])
        m['confirmed_navs_desc'] = jf.get('confirmed_navs_desc', [])
        m['trend_20d_pct'] = jf.get('trend_20d_pct')
        m['val_metric'] = jf.get('val_metric', '')
        m['val_value'] = jf.get('val_value')
        m['val_signal'] = jf.get('val_signal', '')
        m['val_na_reason'] = jf.get('val_na_reason', '')
        m['tp_display'] = jf.get('tp_display')
        merged.append(m)
    return merged

def fmt_pct(v, fallback='—'):
    if v is None:
        return f'<span class="text-gray">{fallback}</span>'
    try:
        vf = float(v)
    except (ValueError, TypeError):
        return f'<span class="text-gray">{fallback}</span>'
    pct = vf * 100 if abs(vf) < 1 else vf
    sign = '+' if pct >= 0 else ''
    cls = 'up' if pct > 0 else ('down' if pct < 0 else 'flat')
    return f'<span class="pct {cls}">{sign}{pct:.2f}%</span>'

def fmt_daily_change(dc, final_nav_mode=False):
    if dc is None:
        label = '已确认' if final_nav_mode else '休市'
        return f'<span class="pct flat">{label}</span>'
    try:
        v = float(dc)
    except (ValueError, TypeError):
        return '<span class="pct flat">休市</span>'
    # F列在Excel中是小数（如 0.0056 表示 0.56%），data_only=True读出来也是小数，需要乘100
    pct = v * 100 if abs(v) < 1 else v
    sign = '+' if pct >= 0 else ''
    cls = 'up' if pct > 0 else ('down' if pct < 0 else 'flat')
    return f'<span class="pct {cls}">{sign}{pct:.2f}%</span>'

def rsi_line_chart(rsi_history, rsi_current):
    """RSI历史趋势折线图（SVG），显示RSI曲线 + 参考线30/70"""
    if not rsi_history or len(rsi_history) < 2:
        # 回退到横向进度条
        if rsi_current is None:
            return '—'
        try:
            v = float(rsi_current)
        except (ValueError, TypeError):
            return '—'
        pos = max(0, min(100, v))
        cx = pos / 100 * 52 + 2
        if v < 30:
            color = '#27ae60'
        elif v > 70:
            color = '#e74c3c'
        else:
            color = '#f39c12'
        return f'''<div class="rsi-wrap">
      <svg width="56" height="12" style="display:block;margin:0 auto;">
        <rect x="1" y="4" width="54" height="4" fill="#e8e8e8" rx="2"/>
        <rect x="1" y="4" width="{pos/100*54:.1f}" height="4" fill="{color}" rx="2"/>
        <circle cx="{cx:.1f}" cy="6" r="2.5" fill="#333"/>
      </svg>
    </div>'''
    
    data = rsi_history  # oldest→newest
    
    # SVG dimensions
    w, h = 88, 34
    pad_left, pad_right = 4, 2
    pad_top, pad_bottom = 4, 4
    pw = w - pad_left - pad_right  # plot width
    ph = h - pad_top - pad_bottom  # plot height
    
    def rsi_to_y(v):
        """RSI 0→bottom, 100→top"""
        return pad_top + ph - (v / 100.0 * ph)
    
    n = len(data)
    xs = [pad_left + i * pw / (n - 1) for i in range(n)]
    ys = [rsi_to_y(v) for v in data]
    
    # Build polyline points
    points = ' '.join(f'{x:.1f},{y:.1f}' for x, y in zip(xs, ys))
    
    # Current RSI determines line color
    if rsi_current is not None:
        try:
            cv = float(rsi_current)
        except (ValueError, TypeError):
            cv = 50
    else:
        cv = 50
    
    if cv < 30:
        line_color = '#27ae60'
        num_color = '#27ae60'
    elif cv > 70:
        line_color = '#e74c3c'
        num_color = '#e74c3c'
    else:
        line_color = '#4a90d9'
        num_color = '#666'
    
    y30 = rsi_to_y(30)
    y70 = rsi_to_y(70)
    y50 = rsi_to_y(50)
    
    return f'''<div class="rsi-wrap">
      <svg width="{w}" height="{h}" style="display:block;margin:0 auto;">
        <!-- 背景区域 -->
        <rect x="{pad_left}" y="{pad_top}" width="{pw:.1f}" height="{ph:.1f}" fill="#f8f9fa" rx="1"/>
        <!-- 超卖区(0-30) 浅绿 -->
        <rect x="{pad_left}" y="{y30:.1f}" width="{pw:.1f}" height="{ph-y30+pad_top:.1f}" fill="#e8f5e9" opacity="0.7"/>
        <!-- 超买区(70-100) 浅红 -->
        <rect x="{pad_left}" y="{pad_top}" width="{pw:.1f}" height="{y70-pad_top:.1f}" fill="#ffebee" opacity="0.7"/>
        <!-- 参考线 30 -->
        <line x1="{pad_left}" y1="{y30:.1f}" x2="{pad_left+pw:.1f}" y2="{y30:.1f}" stroke="#bdbdbd" stroke-width="0.5" stroke-dasharray="2,2"/>
        <!-- 参考线 50 -->
        <line x1="{pad_left}" y1="{y50:.1f}" x2="{pad_left+pw:.1f}" y2="{y50:.1f}" stroke="#e0e0e0" stroke-width="0.3"/>
        <!-- 参考线 70 -->
        <line x1="{pad_left}" y1="{y70:.1f}" x2="{pad_left+pw:.1f}" y2="{y70:.1f}" stroke="#bdbdbd" stroke-width="0.5" stroke-dasharray="2,2"/>
        <!-- RSI曲线 -->
        <polyline points="{points}" fill="none" stroke="{line_color}" stroke-width="1.3" stroke-linejoin="round" stroke-linecap="round"/>
        <!-- 最后一个点 -->
        <circle cx="{xs[-1]:.1f}" cy="{ys[-1]:.1f}" r="1.8" fill="{line_color}" stroke="#fff" stroke-width="0.5"/>
      </svg>
    </div>'''

def level_badge(lv):
    """M(倍投等级) badge，如：倍投3级"""
    if not lv:
        return '—'
    lv_cls_map = {
        "停止区": "level-1", "观望区": "level-1",
        "倍投1级": "level-1", "倍投2级": "level-2",
        "倍投3级": "level-3", "倍投4级": "level-4",
        "倍投5级": "level-5",
        "补仓1级": "level-1", "补仓2级": "level-2",
        "补仓3级": "level-3", "补仓4级": "level-4",
        "补仓5级": "level-5",
        "极限1级": "level-extreme-1", "极限2级": "level-extreme-2",
        "极限3级": "level-extreme-3",
    }
    cls = lv_cls_map.get(lv, "level-1")
    return f'<span class="level-badge {cls}">{lv}</span>'

def mult_display(mult):
    """N(倍投基数) 单独显示，如：2倍"""
    if mult is None:
        return '—'
    try:
        m = float(mult)
    except (ValueError, TypeError):
        return '—'
    if m == 0:
        return '<span class="text-gray">0倍</span>'
    if m == int(m):
        return f'<span class="mult-val">{int(m)}倍</span>'
    return f'<span class="mult-val">{m:.1f}倍</span>'

def rsi_signal_badge(sig):
    cls_map = {
        "极度超卖": "rsi-extreme", "超卖": "rsi-oversold",
        "偏弱": "rsi-weak", "中性": "rsi-neutral",
        "超买": "rsi-overbought", "N/A": "rsi-na"
    }
    cls = cls_map.get(sig, "rsi-neutral")
    return f'<span class="rsi-badge {cls}">{sig}</span>'

def val_display(f):
    vm = f.get('val_metric', '')
    vv = f.get('val_value')
    vs = f.get('val_signal', '')
    nr = f.get('val_na_reason', '')
    if nr or vm == '不适用':
        raw = f.get('val_display', '') or nr or '不适用'
        return f'<span class="text-gray small">{raw}</span>'
    if vv is None or vm == '':
        raw = f.get('val_display', '') or '—'
        return f'<span class="text-gray small">{raw}</span>'
    label = vs.split()[-1] if ' ' in vs else vs
    if label in ('低估', '跌破净值'):
        color = 'text-green'
    elif label == '适中':
        color = 'text-orange'
    else:
        color = 'text-red'
    return f'{vm}={vv:.1f}<br><span class="{color} small">{label}</span>'

# 操作建议颜色映射 — 文字即颜色，不依赖JSON
ACTION_COLORS = {
    '强烈补仓': ('#1B5E20', '#C8E6C9'),
    '建议补仓': ('#2E7D32', '#C8E6C9'),
    '可补仓':   ('#E65100', '#FFEB9C'),
    '关注':     ('#5D6D7E', '#E8EAF0'),
    '观望':     ('#7F8C8D', '#EAEDED'),
    '暂不操作': ('#95A5A6', '#F0F0F0'),
    '警惕回调': ('#C62828', '#FFCDD2'),
    '考虑止盈': ('#8B0000', '#FFCDD2'),
}

def action_display(f):
    sig = f.get('action', '')
    if not sig:
        return '—'
    tc, bg = ACTION_COLORS.get(sig, ('#333', '#f0f0f0'))
    return f'<span class="action-badge" style="color:{tc};background:{bg};">{sig}</span>'

def tp_display(f):
    tp = f.get('tp_display')
    if tp:
        return f'<span class="tp-text">{tp}</span>'
    raw = f.get('tp_signal')
    if raw:
        return f'<span class="tp-text">{raw}</span>'
    return '<span class="text-gray">—</span>'

def safe_num(v, decimals=0):
    if v is None:
        return '—'
    try:
        n = float(v)
    except (ValueError, TypeError):
        return '—'
    if decimals == 0:
        return f'{n:,.0f}' if n == int(n) else f'{n:.2f}'
    return f'{n:.{decimals}f}'

def safe_mult(v):
    """补仓倍数显示：整数显示整数，小数保留1位"""
    if v is None:
        return '—'
    try:
        n = float(v)
    except (ValueError, TypeError):
        return '—'
    if n == 0:
        return '<span class="text-gray">0</span>'
    txt = f'{n:.0f}' if n == int(n) else f'{n:.1f}'
    return f'<span class="guide-mult">{txt}</span>'

def safe_amt(v):
    """补仓金额显示"""
    if v is None:
        return '—'
    try:
        n = float(v)
    except (ValueError, TypeError):
        return '—'
    if n == 0:
        return '<span class="text-gray">¥0</span>'
    return f'<span class="amt-val">¥{n:,.0f}</span>'

def safe_shares(v):
    if v is None:
        return '—'
    try:
        n = float(v)
    except (ValueError, TypeError):
        return '—'
    return f'<span class="shares-val">{n:.0f}</span>'

def safe_date(v):
    if not v:
        return '—'
    return str(v)[:10]

def safe_nav(v):
    if v is None:
        return '—'
    try:
        n = float(v)
    except (ValueError, TypeError):
        return '—'
    return f'{n:.4f}'

def generate():
    json_data = load_json()
    final_nav_mode = json_data.get('data_mode') == 'final_nav'
    excel_rows = load_excel()
    compute_formulas(excel_rows)  # 在Python中计算G/P/Q/R/T/U/X，不依赖Excel缓存
    funds = merge_data(excel_rows, json_data)
    
    update_time = json_data['update_time']
    date_str = update_time[:10] if update_time else '—'
    regime = json_data.get('market_regime', {})
    
    regime_name = regime.get('name', '—')
    regime_desc = regime.get('desc', '')
    sse_close = regime.get('sse_close')
    sse_ma200 = regime.get('sse_ma200')
    sse_above = regime.get('sse_above_ma200')
    sse_rsi = regime.get('sse_rsi')
    sse_dd = regime.get('sse_drawdown_from_peak')
    sse_change_pct = regime.get('sse_change_pct')
    sse_change_amount = regime.get('sse_change_amount')
    hs300_pct = regime.get('hs300_pe_pct')
    fear_greed = regime.get('fear_greed')
    hk_connect = regime.get('hk_connect')
    margin_balance = regime.get('margin_balance')
    
    buy_signals = ['强烈补仓', '建议补仓', '可补仓']
    need_buy = [f for f in funds if f.get('action', '') in buy_signals]
    need_buy_count = len(need_buy)
    total_shares = sum(
        float(f.get('shares', 0) or 0) for f in need_buy
    )
    total_amount = sum(
        float(f.get('amount', 0) or 0) for f in need_buy
    )
    
    rows = []
    for f in funds:
        row_cls = 'row-watch' if f.get('action') == '关注' else ''
        code = f['code']
        d_val = f.get('d_val', '') or ''
        e_val = f.get('e_val', '') or ''
        o_val = f.get('o_val', '') or ''
        w_val = f.get('trough_nav', '') or ''
        
        action_val = f.get('action', '') or ''
        grade_mult_val = f.get('grade_mult', '') or ''
        unit_price_val = f.get('unit_price', '') or ''
        level_val = f.get('level', '') or ''
        confirmed_navs_json = json.dumps(f.get('confirmed_navs_desc', []), ensure_ascii=False)
        trend_20d_val = f.get('trend_20d_pct')
        val_signal_val = f.get('val_signal', '') or ''
        rows.append(f'''<tr class="{row_cls}" data-code="{code}" data-d="{d_val}" data-e="{e_val}" data-o="{o_val}" data-w="{w_val}" data-action="{action_val}" data-n="{grade_mult_val}" data-s="{unit_price_val}" data-level="{level_val}" data-navs='{confirmed_navs_json}' data-val-signal="{val_signal_val}" data-trend20="{trend_20d_val if trend_20d_val is not None else ''}">
            <td class="col-cat">{f['cat']}</td>
            <td class="col-code">{code}</td>
            <td class="col-name">{f['name']}</td>
            <td class="col-dd down"><span class="live-dd">{fmt_pct(f['drawdown'])}</span></td>
            <td class="col-rsi live-rsi"><div class="rsi-cell-wrap">{rsi_line_chart(f.get('rsi_history', []), f['rsi'])}<div class="rsi-sig-inline">{rsi_signal_badge(f['rsi_signal_j'])}</div></div></td>
            <td class="col-val">{val_display(f)}</td>
            <td class="col-action live-action">{action_display(f)}</td>
            <td class="col-level-combined live-level">{level_badge(f['level'])} {mult_display(f['grade_mult'])}</td>
            <td class="col-change"><span class="live-change">{fmt_daily_change(f['daily_change'], final_nav_mode)}</span></td>
            <td class="col-q"><span class="live-q">{fmt_pct(f['q_val'])}</span></td>
            <td class="col-shares"><span class="live-shares">{safe_shares(f['shares'])}</span></td>
            <td class="col-price">{safe_num(f['unit_price'], 0)}</td>
            <td class="col-final-mult"><span class="live-mult">{safe_mult(f['final_mult'])}</span></td>
            <td class="col-amount"><span class="live-amount">{safe_amt(f['amount'])}</span></td>
            <td class="col-date">{safe_date(f['trough_date'])}</td>
            <td class="col-nav">{safe_nav(f['trough_nav'])}</td>
            <td class="col-x"><span class="live-x">{fmt_pct(f['x_gain'])}</span></td>
            <td class="col-tp">{tp_display(f)}</td>
        </tr>''')
    
    rows_str = '\n'.join(rows)
    
    market_items = []
    market_items.append(f'<div class="m-item"><span class="m-label">市场状态</span><span class="m-val">{regime_name}</span></div>')
    if sse_close:
        change_html = ''
        if sse_change_pct is not None and sse_change_amount is not None:
            if sse_change_pct > 0:
                change_html = f'<span class="m-change up">(较前日+{sse_change_pct:.2f}% +{sse_change_amount:.2f})</span>'
            elif sse_change_pct < 0:
                change_html = f'<span class="m-change down">(较前日{sse_change_pct:.2f}% {sse_change_amount:.2f})</span>'
            else:
                change_html = f'<span class="m-change flat">(较前日0.00%)</span>'
        market_items.append(f'<div class="m-item"><span class="m-label">上证指数</span><span class="m-val" id="live-sse">{sse_close:.0f}{change_html}</span></div>')
    if sse_ma200:
        market_items.append(f'<div class="m-item"><span class="m-label">200日线</span><span class="m-val">{sse_ma200:.0f}</span></div>')
    if sse_rsi:
        market_items.append(f'<div class="m-item"><span class="m-label">上证RSI</span><span class="m-val">{sse_rsi:.1f}</span></div>')
    if sse_dd is not None:
        # 根据上证当天涨跌变换颜色：涨=红色，跌=绿色
        dd_color = ''
        if sse_change_pct is not None:
            if sse_change_pct > 0:
                dd_color = ' style="color:#E74C3C"'
            elif sse_change_pct < 0:
                dd_color = ' style="color:#27AE60"'
        market_items.append(f'<div class="m-item"><span class="m-label">峰值回撤</span><span class="m-val"{dd_color}>{sse_dd:.1f}%</span></div>')
    # 恐贪指数（按韭圈儿颜色标准：恐惧=绿色，贪婪=红色）
    if fear_greed:
        fg_val = fear_greed.get('value')
        fg_label = fear_greed.get('label', '')
        fg_date = fear_greed.get('date', '')
        # 日期只显示月-日
        if len(fg_date) >= 10:
            fg_date = fg_date[5:]
        # 韭圈儿颜色标准
        if fg_label == '极度恐惧':
            fg_cls = 'fg-extreme-fear'
            fg_color = '#1E3F1F'
        elif fg_label == '恐惧':
            fg_cls = 'fg-fear'
            fg_color = '#5C9A58'
        elif fg_label in ('中立', '中性'):
            fg_cls = 'fg-neutral'
            fg_color = '#3B6FA8'
        elif fg_label == '贪婪':
            fg_cls = 'fg-greed'
            fg_color = '#B86060'
        elif fg_label == '极度贪婪':
            fg_cls = 'fg-extreme-greed'
            fg_color = '#6B2020'
        else:
            fg_cls = 'fg-neutral'
            fg_color = '#85929E'
        market_items.append(f'<div class="m-item"><span class="m-label">恐贪指数({fg_date})</span><span class="m-val"><span class="{fg_cls}" style="color:{fg_color}; font-weight:700;">{fg_val}</span> <span class="fg-badge {fg_cls}">{fg_label}</span></span></div>')
    
    # 北向/南向资金
    if hk_connect and hk_connect.get('north_net') is not None:
        hk_date = hk_connect.get('date', '')
        if len(hk_date) >= 10:
            hk_date = hk_date[5:]
        north = hk_connect['north_net']
        south = hk_connect['south_net']
        north_cls = 'up' if north > 0 else 'down' if north < 0 else 'flat'
        south_cls = 'up' if south > 0 else 'down' if south < 0 else 'flat'
        market_items.append(f'<div class="m-item"><span class="m-label">北向({hk_date})</span><span class="m-val"><span class="m-change {north_cls}">{north:+.2f}亿</span></span></div>')
        market_items.append(f'<div class="m-item"><span class="m-label">南向({hk_date})</span><span class="m-val"><span class="m-change {south_cls}">{south:+.2f}亿</span></span></div>')
    
    # 融资余额（与基金补仓自动化一致：加杠杆/去杠杆标签）
    if margin_balance:
        m_date = margin_balance.get('date', '')
        if len(m_date) >= 10:
            m_date = m_date[5:]
        m_balance = margin_balance.get('balance')  # 单位：亿元
        m_net_buy = margin_balance.get('net_buy')
        m_state = margin_balance.get('state', '')
        m_meaning = margin_balance.get('meaning', '')
        m_state_color = margin_balance.get('state_color', '#7F8C8D')
        # 余额：亿元 → 万亿元，保留2位小数
        m_balance_trillion = round(m_balance / 1e4, 2) if m_balance is not None else None
        balance_str = f'{m_balance_trillion:.2f}万亿元' if m_balance_trillion is not None else '—'
        # 净买入金额（红增绿减）
        if m_net_buy is not None:
            m_cls = 'up' if m_net_buy > 0 else 'down' if m_net_buy < 0 else 'flat'
            net_buy_str = f'<span class="m-change {m_cls}">{m_net_buy:+.0f}亿</span>'
        else:
            net_buy_str = ''
        # 状态标签
        tag_str = f'<span style="color:{m_state_color}; font-weight:700;">（{m_state}；{m_meaning}）</span>' if m_state and m_meaning else ''
        market_items.append(f'<div class="m-item"><span class="m-label">融资余额({m_date})</span><span class="m-val">{balance_str} {net_buy_str}{tag_str}</span></div>')
    
    market_html = '\n'.join(market_items)
    
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>金字塔丛林补仓指导图 ({date_str})</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{
  font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
  background: #eef2f5;
  color: #333;
  line-height: 1.5;
  padding: 16px;
  font-size: 13px;
}}
.container {{ max-width: 1800px; margin: 0 auto; }}

.top-bar {{
  background: linear-gradient(135deg, #1e3a5f 0%, #2c5282 100%);
  color: #fff;
  padding: 14px 24px;
  border-radius: 10px 10px 0 0;
  display: flex;
  align-items: center;
  justify-content: space-between;
  flex-wrap: wrap;
  gap: 8px;
}}
.top-bar .main-title {{ font-size: 18px; font-weight: bold; letter-spacing: 1px; }}
.top-bar .stats {{ font-size: 14px; opacity: 0.95; }}
.top-bar .highlight {{ color: #ffd700; font-weight: bold; }}

.market-bar {{
  background: #e8f0f8;
  padding: 10px 24px;
  display: flex;
  flex-wrap: wrap;
  gap: 20px;
  align-items: center;
  border-bottom: 2px solid #d0dce8;
}}
.market-bar .m-item {{ display: flex; align-items: baseline; gap: 6px; }}
.market-bar .m-label {{ font-size: 14px; color: #000; font-weight: bold; }}
.market-bar .m-val {{ font-size: 13px; font-weight: bold; color: #1a365d; }}
.market-bar .m-change {{ font-size: 12px; font-weight: normal; margin-left: 4px; }}
.market-bar .m-change.up {{ color: #c62828; }}
.market-bar .m-change.down {{ color: #2e7d32; }}
.market-bar .m-change.flat {{ color: #7f8c8d; }}
.market-bar .fg-badge {{ display: inline-block; padding: 1px 6px; border-radius: 10px; font-size: 11px; font-weight: normal; margin-left: 4px; color: #fff; }}
.market-bar .fg-badge.fg-extreme-fear {{ background: #1E3F1F; }}
.market-bar .fg-badge.fg-fear {{ background: #5C9A58; }}
.market-bar .fg-badge.fg-neutral {{ background: #3B6FA8; }}
.market-bar .fg-badge.fg-greed {{ background: #B86060; }}
.market-bar .fg-badge.fg-extreme-greed {{ background: #6B2020; }}

.table-wrap {{
  background: #fff;
  border-radius: 0 0 10px 10px;
  overflow-x: auto;
  box-shadow: 0 4px 16px rgba(0,0,0,0.08);
}}
table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 12.5px;
  min-width: 1500px;
}}

thead {{ background: #2c3e50; color: #fff; }}
.region-row th {{
  padding: 7px 4px;
  font-size: 14px;
  font-weight: bold;
  text-align: center;
  border-right: 1px solid rgba(255,255,255,0.10);
}}
.region-info {{ background: #556572; color: #fff; }}
.region-indicators {{ background: #526B58; color: #fff; }}
.region-buy {{ background: #8B7355; color: #fff; }}
.region-tp {{ background: #DBA9B5; color: #fff; }}

.header-row th {{
  padding: 8px 4px;
  font-size: 12px;
  font-weight: 600;
  text-align: center;
  white-space: nowrap;
  border-right: 1px solid rgba(255,255,255,0.10);
  line-height: 1.4;
}}
.header-row th:last-child {{ border-right: none; }}

/* 版块分隔线：加深版块边界的视觉区分 */
col.section-divider {{ border-left: 2.5px solid rgba(0,0,0,0.10); }}

tbody tr {{ border-bottom: 2px solid #d5dae0; transition: background 0.12s; }}
tbody tr:hover {{ background: rgba(0,0,0,0.03); }}
tbody tr.row-watch {{ background: #FFFBEA; }}
tbody tr.row-watch:hover {{ background: #FFF6D6; }}

tbody td {{
  padding: 7px 4px;
  text-align: center;
  font-size: 12px;
  border-right: 1px solid rgba(0,0,0,0.04);
  vertical-align: middle;
}}
tbody td:last-child {{ border-right: none; }}

.col-cat {{ min-width: 72px; font-weight: 500; color: #555; font-size: 11.5px; }}
.col-code {{ min-width: 60px; color: #7f8c8d; font-size: 12px; font-family: monospace; }}
.col-name {{ min-width: 165px; text-align: left; padding-left: 8px; font-weight: 500; color: #2c3e50; font-size: 14px; }}
.col-dd {{ min-width: 68px; font-weight: bold; }}
.col-dd .pct {{ font-size: 13px; }}
.col-rsi {{ min-width: 140px; }}
.col-val {{ min-width: 95px; line-height: 1.4; font-size: 11.5px; }}
.col-action {{ min-width: 88px; }}
.col-level-combined {{ min-width: 115px; white-space: nowrap; }}
.col-change {{ min-width: 68px; }}
.col-q {{ min-width: 78px; }}
.col-shares {{ min-width: 52px; }}
.col-price {{ min-width: 48px; color: #666; }}
.col-final-mult {{ min-width: 60px; }}
.col-amount {{ min-width: 72px; }}
.col-date {{ min-width: 82px; font-size: 11px; }}
.col-nav {{ min-width: 72px; font-family: monospace; }}
.col-x {{ min-width: 72px; }}
.col-x .pct {{ font-size: 14px; }}
.col-tp {{ min-width: 95px; }}

.pct {{ font-weight: bold; font-family: monospace; font-size: 13px; }}
.pct.up {{ color: #e74c3c; }}
.pct.down {{ color: #27ae60; }}
.pct.flat {{ color: #95a5a6; }}

.text-green {{ color: #27ae60; font-weight: bold; }}
.text-red {{ color: #e74c3c; font-weight: bold; }}
.text-orange {{ color: #e67e22; font-weight: bold; }}
.text-gray {{ color: #95a5a6; }}
.small {{ font-size: 11px; }}

.level-badge {{
  display: inline-block;
  padding: 2px 8px;
  border-radius: 10px;
  font-size: 11px;
  font-weight: bold;
  white-space: nowrap;
}}
.level-extreme-3 {{ background: #fadbd8; color: #922b21; border: 1px solid #c0392b; }}
.level-extreme-2 {{ background: #f5b7b1; color: #922b21; border: 1px solid #c0392b; }}
.level-extreme-1 {{ background: #fdf2f2; color: #c0392b; border: 1px solid #e74c3c; }}
.level-5 {{ background: #fef5e7; color: #d35400; border: 1px solid #e67e22; }}
.level-4 {{ background: #fef9e7; color: #b7950b; border: 1px solid #f1c40f; }}
.level-3 {{ background: #eafaf1; color: #1e8449; border: 1px solid #2ecc71; }}
.level-2 {{ background: #ebf5fb; color: #2471a3; border: 1px solid #3498db; }}
.level-1 {{ background: #f2f3f4; color: #7f8c8d; border: 1px solid #95a5a6; }}

.level-mult {{ font-size: 11px; opacity: 0.85; }}

.rsi-badge {{
  display: inline-block;
  padding: 2px 7px;
  border-radius: 8px;
  font-size: 11px;
  font-weight: bold;
  white-space: nowrap;
}}
.rsi-extreme {{ background: #1b5e20; color: #fff; }}
.rsi-oversold {{ background: #e8f5e9; color: #2e7d32; border: 1px solid #a5d6a7; }}
.rsi-weak {{ color: #3498db; }}
.rsi-neutral {{ color: #666; }}
.rsi-overbought {{ background: #ffebee; color: #c62828; border: 1px solid #ef9a9a; }}

.rsi-wrap {{ display: flex; flex-direction: column; align-items: center; gap: 2px; }}
.rsi-cell-wrap {{ display: flex; flex-direction: row; align-items: center; gap: 5px; }}
.rsi-sig-inline {{ flex-shrink: 0; }}

.action-badge {{
  display: inline-block;
  padding: 3px 8px;
  border-radius: 6px;
  font-size: 11.5px;
  font-weight: bold;
  white-space: nowrap;
}}

.mult-val {{ font-weight: bold; color: #1a5276; font-size: 13px; }}
.guide-mult {{ font-weight: bold; color: #1a5276; font-size: 15px; }}
.shares-val {{ font-weight: bold; color: #1b5e20; font-size: 13px; }}
.amt-val {{ font-weight: bold; color: #333; font-size: 14px; }}

.tp-text {{ color: #8B0000; font-weight: bold; font-size: 11.5px; }}

.footer {{
  margin-top: 12px;
  padding: 10px 16px;
  background: #fff;
  border-radius: 8px;
  font-size: 11.5px;
  color: #666;
  line-height: 1.8;
  box-shadow: 0 1px 4px rgba(0,0,0,0.04);
}}
.footer strong {{ color: #444; }}

/* ===== 手机横屏增强：保留完整表格的一览式框架 ===== */
.mobile-orientation-bar {{
  display: none;
  background: #173a62;
  color: #fff;
  padding: 9px 12px;
  border-radius: 10px 10px 0 0;
  align-items: center;
  justify-content: space-between;
  gap: 10px;
  box-shadow: 0 2px 8px rgba(0,0,0,.12);
}}
.orientation-copy {{ min-width: 0; }}
.orientation-title {{ font-size: 13px; font-weight: 700; }}
.orientation-hint {{ font-size: 10px; opacity: .84; margin-top: 1px; }}
.orientation-actions {{ display: flex; align-items: center; gap: 6px; flex-shrink: 0; }}
.orientation-btn {{
  appearance: none; border: 1px solid rgba(255,255,255,.42); background: #fff;
  color: #173a62; border-radius: 7px; padding: 7px 9px; font-weight: 700;
  font-size: 12px; white-space: nowrap; cursor: pointer;
}}
.orientation-btn.secondary {{ background: transparent; color: #fff; display: none; }}
.rotate-tip {{
  display: none; margin: 0; padding: 8px 12px; background: #fff6df;
  color: #795700; border: 1px solid #f1d890; border-top: 0; font-size: 11px;
}}
body.preview-landscape .orientation-btn.secondary {{ display: inline-block; }}
body.preview-landscape .top-bar {{ border-radius: 0; }}
/* 以高度识别横屏手机，避免横屏后 CSS 宽度超过 768px 而误回桌面样式。 */
@media (max-width: 768px), (max-height: 500px) and (orientation: landscape) {{
  body {{ padding: 6px; }}
  .container {{ max-width: none; }}
  .mobile-orientation-bar {{ display: flex; }}
  .top-bar {{ padding: 9px 12px; gap: 4px; flex-direction: column; align-items: flex-start; }}
  .top-bar .main-title {{ font-size: 15px; letter-spacing: .3px; }}
  .top-bar .stats {{ display: block; margin-top: 3px; font-size: 12px; }}
  .market-bar {{ padding: 7px 10px; gap: 8px 13px; max-height: 76px; overflow: hidden; }}
  .market-bar .m-label {{ font-size: 11px; }}
  .market-bar .m-val {{ font-size: 11px; }}
  .footer {{ margin-top: 8px; padding: 8px 10px; font-size: 10px; }}
  .table-wrap {{ border-radius: 0 0 8px 8px; -webkit-overflow-scrolling: touch; }}
}}
@media ((max-width: 768px) and (orientation: landscape)), ((max-height: 500px) and (orientation: landscape)) {{
  body {{ padding: 4px; font-size: 11px; }}
  .mobile-orientation-bar {{ padding: 5px 9px; border-radius: 7px 7px 0 0; }}
  .orientation-title {{ font-size: 11px; }}
  .orientation-hint {{ display: none; }}
  .orientation-btn {{ font-size: 10px; padding: 4px 7px; }}
  .top-bar {{ padding: 6px 10px; min-height: 34px; flex-direction: row; align-items: center; }}
  .top-bar .main-title {{ font-size: 13px; }}
  .top-bar .stats {{ display: inline; margin-left: 6px; font-size: 10px; }}
  .top-bar > div:last-child {{ font-size: 9px !important; }}
  .market-bar {{ padding: 5px 8px; gap: 4px 11px; max-height: 48px; overflow: auto; }}
  .market-bar .m-label, .market-bar .m-val {{ font-size: 10px; }}
  .table-wrap {{ max-height: calc(100vh - 98px); overflow: auto; isolation: isolate; }}
  table {{ min-width: 1350px; font-size: 10.5px; border-collapse: separate; border-spacing: 0; }}
  .region-row th {{ padding: 4px 3px; font-size: 11px; height: 26px; }}
  .header-row th {{ padding: 5px 3px; font-size: 10px; height: 30px; }}
  tbody td {{ padding: 4px 3px; font-size: 10px; }}
  .col-cat {{ min-width: 58px; font-size: 10px; }}
  .col-code {{ min-width: 53px; font-size: 10px; }}
  .col-name {{ min-width: 130px; font-size: 11px; padding-left: 5px; }}
  .col-rsi {{ min-width: 116px; }}
  .col-val {{ min-width: 80px; font-size: 10px; }}
  .col-action {{ min-width: 76px; }}
  .col-level-combined {{ min-width: 94px; }}
  .rsi-badge, .level-badge, .action-badge {{ font-size: 9.5px; padding: 2px 5px; }}
  .guide-mult, .amt-val, .pct, .pct.up, .pct.down {{ font-size: 11px; }}
  /* 两级标题固定在表格滚动区顶部，避免数据行穿透覆盖。 */
  thead {{ position: relative; z-index: 40; }}
  thead .region-row th {{ position: sticky; top: 0; z-index: 42; background-clip: padding-box; box-shadow: 0 1px 0 rgba(255,255,255,.2); }}
  thead .header-row th {{ position: sticky; top: 26px; z-index: 41; background: #2c3e50; background-clip: padding-box; box-shadow: 0 2px 3px rgba(0,0,0,.22); }}
  /* 横向滚动时固定基金上下文。 */
  thead .header-row th:nth-child(1), tbody td:nth-child(1) {{ position: sticky; left: 0; z-index: 20; }}
  thead .header-row th:nth-child(2), tbody td:nth-child(2) {{ position: sticky; left: 58px; z-index: 20; }}
  thead .header-row th:nth-child(3), tbody td:nth-child(3) {{ position: sticky; left: 111px; z-index: 20; box-shadow: 3px 0 5px rgba(0,0,0,.13); }}
  .region-row th.region-info {{ position: sticky; left: 0; z-index: 45; }}
  thead .header-row th:nth-child(-n+3) {{ background: #556572; z-index: 46; }}
  thead .region-row th.region-info {{ z-index: 47; }}
  tbody td:nth-child(1), tbody td:nth-child(2), tbody td:nth-child(3) {{ background: #E8EDF2; }}
  tbody tr.row-watch td:nth-child(1), tbody tr.row-watch td:nth-child(2), tbody tr.row-watch td:nth-child(3) {{ background: #FFFBEA; }}
  .footer {{ display: none; }}
}}
/* 实时/收盘状态独占第二行，不覆盖涨跌数字原有颜色、字号或粗细。 */
.live-est {{
  display: inline-flex;
  flex-direction: column;
  align-items: center;
  line-height: 1.15;
}}
.live-est .live-tag {{
  display: block;
  margin-top: 2px;
  font-size: 9px;
  color: #95a5a6;
  font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
  font-weight: normal;
  line-height: 1;
  white-space: nowrap;
}}
.live-indicator {{
  display: inline-block;
  width: 5px;
  height: 5px;
  border-radius: 50%;
  background: #95a5a6;
  margin-right: 2px;
  vertical-align: 1px;
}}
.live-indicator.is-live {{ animation: live-pulse 2s infinite; }}
@keyframes live-pulse {{
  0%, 100% {{ opacity: 1; }}
  50% {{ opacity: 0.3; }}
}}
</style>
<script>
var FINAL_NAV_MODE = {str(final_nav_mode).lower()};
// ========== 路线B 实时数据引擎 v2.0 ==========
// 功能：盘中从天天基金API实时拉取净值 → 重算G/Q/R/S/T/U/U总计 → 更新页面
// 时段：9:30-11:30/13:00-14:29 每5分钟 | 14:30-15:00 每30秒 | 盘后60秒整页刷新
var LIVE_INTERVAL = 30 * 1000;     // 30秒（14:30-15:00高频窗口）
var SLOW_INTERVAL = 5 * 60 * 1000; // 5分钟（盘中慢速）
var COUNTDOWN_START = 60;          // 页面刷新倒计时

// 补仓信号关键词
var BUY_SIGNALS = ['强烈补仓', '建议补仓', '可补仓'];

// 从表格行解析基金数据（含完整数据属性）
var funds = [];
function initFunds() {{
  if (funds.length) return;
  // HTML 脚本位于 head，需在 DOM 构建完成后再读取基金行。
  var rows = document.querySelectorAll('tbody tr');
  rows.forEach(function(row) {{
    var code = row.getAttribute('data-code');
    if (!code) return;
    funds.push({{
      code: code,
      row: row,
      peak: parseFloat(row.getAttribute('data-d')) || 0,
      nav: parseFloat(row.getAttribute('data-e')) || 0,
      anchor: parseFloat(row.getAttribute('data-o')) || 0,
      trough: parseFloat(row.getAttribute('data-w')) || 0,
      action: row.getAttribute('data-action') || '',
      grade_mult: parseFloat(row.getAttribute('data-n')) || 0,
      unit_price: parseFloat(row.getAttribute('data-s')) || 0,
      level: row.getAttribute('data-level') || '',
      confirmed_navs_desc: (function() {{
        try {{ return JSON.parse(row.getAttribute('data-navs') || '[]'); }} catch (e) {{ return []; }}
      }})(),
      val_signal: row.getAttribute('data-val-signal') || '',
      trend_20d_pct: parseFloat(row.getAttribute('data-trend20')),
      staticCells: {{
        change: (row.querySelector('.live-change') || {{}}).innerHTML || '',
        dd: (row.querySelector('.live-dd') || {{}}).innerHTML || '',
        rsi: (row.querySelector('.live-rsi') || {{}}).innerHTML || '',
        action: (row.querySelector('.live-action') || {{}}).innerHTML || '',
        level: (row.querySelector('.live-level') || {{}}).innerHTML || '',
        q: (row.querySelector('.live-q') || {{}}).innerHTML || '',
        shares: (row.querySelector('.live-shares') || {{}}).innerHTML || '',
        mult: (row.querySelector('.live-mult') || {{}}).innerHTML || '',
        amount: (row.querySelector('.live-amount') || {{}}).innerHTML || '',
        x: (row.querySelector('.live-x') || {{}}).innerHTML || ''
      }}
    }});
  }});
}}

// 实时估值由 GitHub Actions 抓取并发布为本站同源快照，避免浏览器跨域请求被数据源拒绝。
var LIVE_SNAPSHOT_URL = 'live_estimates.json';
var liveSnapshotPromise = null;
function loadLiveSnapshot() {{
  if (liveSnapshotPromise) return liveSnapshotPromise;
  liveSnapshotPromise = fetch(LIVE_SNAPSHOT_URL + '?_=' + Date.now(), {{ cache: 'no-store' }})
    .then(function(response) {{
      if (!response.ok) throw new Error('snapshot unavailable');
      return response.json();
    }})
    .then(function(snapshot) {{
      var now = new Date();
      var today = now.getFullYear() + '-' + String(now.getMonth() + 1).padStart(2, '0') + '-' + String(now.getDate()).padStart(2, '0');
      if (!snapshot || snapshot.trade_date !== today || !snapshot.estimates) return null;
      return snapshot;
    }})
    .catch(function() {{ return null; }});
  return liveSnapshotPromise;
}}

function fetchEstimate(code) {{
  return loadLiveSnapshot().then(function(snapshot) {{
    return snapshot && snapshot.estimates ? snapshot.estimates[code] || null : null;
  }});
}}

function updateSseIndex(snapshot) {{
  if (!snapshot || !snapshot.sse_index) return;
  var sse = snapshot.sse_index;
  var target = document.getElementById('live-sse');
  if (!target || !isFinite(sse.close) || !isFinite(sse.change_pct)) return;
  var cls = sse.change_pct > 0 ? 'up' : (sse.change_pct < 0 ? 'down' : 'flat');
  var pctSign = sse.change_pct > 0 ? '+' : '';
  var changeSign = sse.change > 0 ? '+' : '';
  target.innerHTML = Number(sse.close).toFixed(0)
    + '<span class="m-change ' + cls + '">(较前日' + pctSign + Number(sse.change_pct).toFixed(2)
    + '% ' + changeSign + Number(sse.change).toFixed(2) + ')</span>';
}}

function roundTo(value, decimals) {{
  var factor = Math.pow(10, decimals);
  return Math.round((value + Number.EPSILON) * factor) / factor;
}}

function getRsiSignal(rsi) {{
  if (!isFinite(rsi)) return 'N/A';
  if (rsi < 20) return '极度超卖';
  if (rsi < 30) return '超卖';
  if (rsi < 45) return '偏弱';
  if (rsi <= 70) return '中性';
  return '超买';
}}

function calcRsiWithToday(confirmedNavsDesc, estimatedNav, period) {{
  period = period || 14;
  var navs = [estimatedNav].concat(confirmedNavsDesc || []).slice(0, period + 1);
  if (navs.length < period + 1) return null;
  var gains = 0, losses = 0;
  for (var i = 0; i < period; i++) {{
    var change = navs[i] - navs[i + 1];
    if (change > 0) gains += change;
    if (change < 0) losses -= change;
  }}
  if (losses === 0) return 100;
  return roundTo(100 - 100 / (1 + (gains / period) / (losses / period)), 1);
}}

function getLiveLevel(drawdownPct) {{
  var levels = [
    ['停止区', 0, 13, 0], ['观望区', 13, 15, 0.5], ['倍投1级', 15, 20, 1],
    ['倍投2级', 20, 25, 1.5], ['倍投3级', 25, 33, 2], ['倍投4级', 33, 40, 3],
    ['倍投5级', 40, 48, 4], ['极限1级', 48, 55, 5], ['极限2级', 55, 65, 6],
    ['极限3级', 65, Infinity, 7]
  ];
  for (var i = 0; i < levels.length; i++) {{
    if (drawdownPct >= levels[i][1] && drawdownPct < levels[i][2]) {{
      return {{ name: levels[i][0], multiplier: levels[i][3] }};
    }}
  }}
  return {{ name: '极限3级', multiplier: 7 }};
}}

function valuationClass(valSignal) {{
  var value = String(valSignal || '');
  if (value.indexOf('偏低') >= 0 || value.indexOf('低估') >= 0 || value.indexOf('跌破净值') >= 0 || value.indexOf('适中') >= 0) return 'good';
  if (value.indexOf('偏高') >= 0 || value.indexOf('高估') >= 0) return 'bad';
  return 'na';
}}

function getLiveAdvice(drawdownPct, rsi, valSignal, trend20dPct, regimeParams) {{
  var valCls = valuationClass(valSignal);
  var recovering = isFinite(trend20dPct) && trend20dPct > 0;
  var strong = regimeParams && isFinite(regimeParams.rsi_tp_strong) ? regimeParams.rsi_tp_strong : 75;
  var valThreshold = regimeParams && isFinite(regimeParams.rsi_tp_val) ? regimeParams.rsi_tp_val : 70;
  var oversold = isFinite(rsi) && rsi < 30;
  var overbought = isFinite(rsi) && rsi > valThreshold;
  if (isFinite(rsi) && rsi > strong) return {{ action: '考虑止盈', meaning: 'RSI超高位，考虑分批止盈' }};
  if (isFinite(rsi) && rsi > valThreshold && valCls === 'bad') return {{ action: '考虑止盈', meaning: '估值偏高+RSI超买，考虑止盈' }};
  if (drawdownPct < 13) return {{ action: '暂不操作', meaning: '趋势完好，无需补仓' }};
  if (drawdownPct < 15) return {{ action: '暂不操作', meaning: '接近警戒线，可关注等待' }};
  if (drawdownPct < 20) return {{ action: oversold ? '观望' : '观望', meaning: oversold ? '回撤尚浅，等跌幅加深' : '信号不充分，持续观察' }};
  if (drawdownPct < 33) {{
    if (overbought) return {{ action: '警惕回调', meaning: '价格已反弹，不宜追高' }};
    if (oversold) return {{ action: recovering ? '关注' : '可补仓', meaning: '回撤达标+RSI超卖，择机补仓' }};
    return {{ action: '关注', meaning: '回撤达标但RSI未超卖，等待' }};
  }}
  if (drawdownPct < 55) {{
    if (overbought) return {{ action: '警惕回调', meaning: '深跌后已反弹，不宜追高' }};
    if (oversold) {{
      if (valCls === 'bad') return {{ action: recovering ? '关注' : '可补仓', meaning: '回撤深+RSI超卖，但估值偏高需谨慎' }};
      if (valCls === 'good') return {{ action: recovering ? '建议补仓' : '强烈补仓', meaning: '三指标共振，强烈建议补仓' }};
      return {{ action: recovering ? '可补仓' : '建议补仓', meaning: '回撤深+RSI超卖' }};
    }}
    if (valCls === 'bad') return {{ action: '关注', meaning: '回撤深但估值高+RSI中性，等待确认' }};
    return {{ action: recovering ? '关注' : '可补仓', meaning: '回撤深但RSI未超卖，择机补仓' }};
  }}
  if (overbought) return {{ action: '警惕回调', meaning: '深跌后已大幅反弹，不宜追高' }};
  if (oversold) {{
    if (valCls === 'bad') return {{ action: recovering ? '可补仓' : '建议补仓', meaning: '回撤极深+RSI超卖，但估值偏高需谨慎' }};
    if (valCls === 'good') return {{ action: recovering ? '建议补仓' : '强烈补仓', meaning: '三指标共振，历史级底部机会' }};
    return {{ action: recovering ? '建议补仓' : '强烈补仓', meaning: '回撤极深+RSI超卖，历史性底部区域' }};
  }}
  if (valCls === 'bad') return {{ action: recovering ? '关注' : '可补仓', meaning: '回撤极深但估值偏高，谨慎关注' }};
  return {{ action: recovering ? '可补仓' : '建议补仓', meaning: '回撤极深+估值低位，择机补仓' }};
}}

function rsiBadgeClass(signal) {{
  return {{ '极度超卖':'rsi-extreme', '超卖':'rsi-oversold', '偏弱':'rsi-weak', '中性':'rsi-neutral', '超买':'rsi-overbought', 'N/A':'rsi-na' }}[signal] || 'rsi-neutral';
}}
function levelBadgeClass(level) {{
  return {{ '极限3级':'level-extreme-3', '极限2级':'level-extreme-2', '极限1级':'level-extreme-1', '倍投5级':'level-5', '倍投4级':'level-4', '倍投3级':'level-3', '倍投2级':'level-2', '倍投1级':'level-1', '观望区':'level-1', '停止区':'level-1' }}[level] || 'level-1';
}}
function actionColors(action) {{
  return {{ '强烈补仓':['#1B5E20','#C6EFCE'], '建议补仓':['#2E7D32','#C6EFCE'], '可补仓':['#E65100','#FFEB9C'], '关注':['#5D6D7E','#D9E2EC'], '观望':['#7F8C8D','#D9D9D9'], '暂不操作':['#95A5A6','#D9D9D9'], '警惕回调':['#C62828','#FFC7CE'], '考虑止盈':['#8B0000','#FFE0B2'] }}[action] || ['#333','#f0f0f0'];
}}

// 核心：基于实时估值重算盘中临时指标，不写回Excel或确认JSON。
function updateCell(fd, data, snapshotFund, regimeParams) {{
  if (!data || data.gszzl === undefined || data.gszzl === null || data.gszzl === '') return;
  var gszzl = parseFloat(data.gszzl);  // 实时涨跌幅（百分比）
  var row = fd.row;
  var mode = getRefreshMode();
  var isAfterClose = mode === 'afterclose';
  var statusLabel = isAfterClose ? '已收盘' : '实时';
  var indicatorClass = isAfterClose ? '' : ' is-live';

  // 1. F列-当日涨跌（实时替换；状态标签独立显示在第二行）
  var changeCell = row.querySelector('.live-change');
  if (changeCell) {{
    var sign = gszzl >= 0 ? '+' : '';
    var cls = gszzl > 0 ? 'up' : (gszzl < 0 ? 'down' : 'flat');
    changeCell.innerHTML = '<span class="pct ' + cls + ' live-est"><span>' + sign + gszzl.toFixed(2) + '%</span><span class="live-tag"><span class="live-indicator' + indicatorClass + '"></span>' + statusLabel + '</span></span>';
    changeCell.setAttribute('data-live', '1');
  }}

  if (fd.nav <= 0) return;

  // 2. G列-峰值回撤（实时重算）= (Peak - EstNAV) / Peak。
  // 与 Python 同步使用 4 位净值精度和 1 位回撤精度，避免阈值边界出现分歧。
  var estNav = roundTo(fd.nav * (1 + gszzl / 100), 4);
  var newDD = 0;
  if (fd.peak > 0) {{
    newDD = (fd.peak - estNav) / fd.peak * 100;
    var ddCell = row.querySelector('.live-dd');
    if (ddCell) {{
      var ddCls = newDD < 0 ? 'up' : (newDD > 0 ? 'down' : 'flat');
      ddCell.innerHTML = '<span class="pct ' + ddCls + ' live-est">' + newDD.toFixed(2) + '%</span>';
      ddCell.setAttribute('data-live', '1');
    }}
  }}

  // 3. H/I/L/M：用与确认净值脚本相同的规则临时重算，不改变 Excel 静态底稿。
  var liveInputs = snapshotFund || {{}};
  var confirmedNavs = liveInputs.confirmed_navs_desc || fd.confirmed_navs_desc || [];
  var liveRsi = calcRsiWithToday(confirmedNavs, estNav, 14);
  var liveRsiSignal = getRsiSignal(liveRsi);
  var roundedDD = roundTo(newDD, 1);
  var liveLevel = getLiveLevel(roundedDD);
  var liveTrend = isFinite(liveInputs.trend_20d_pct) ? liveInputs.trend_20d_pct : fd.trend_20d_pct;
  var liveValSignal = liveInputs.val_signal || fd.val_signal;
  var liveAdvice = getLiveAdvice(roundedDD, liveRsi, liveValSignal, liveTrend, regimeParams);
  var rsiCell = row.querySelector('.live-rsi');
  if (rsiCell) {{
    rsiCell.innerHTML = '<div class="rsi-cell-wrap"><span class="pct ' + (liveRsi < 30 ? 'down' : (liveRsi > 70 ? 'up' : 'flat')) + ' live-est">'
      + (liveRsi === null ? '—' : liveRsi.toFixed(1)) + '</span><div class="rsi-sig-inline"><span class="rsi-badge ' + rsiBadgeClass(liveRsiSignal) + '">' + liveRsiSignal + '</span></div></div>';
    rsiCell.setAttribute('data-live', '1');
  }}
  var actionCell = row.querySelector('.live-action');
  if (actionCell) {{
    var actionStyle = actionColors(liveAdvice.action);
    actionCell.innerHTML = '<span class="action-badge" style="color:' + actionStyle[0] + ';background:' + actionStyle[1] + ';">' + liveAdvice.action + '</span>';
    actionCell.title = liveAdvice.meaning + (liveTrend > 0 ? '；20日反弹中' : '；20日下跌/横盘');
    actionCell.setAttribute('data-live', '1');
  }}
  var levelCell = row.querySelector('.live-level');
  if (levelCell) {{
    var levelMultText = liveLevel.multiplier === 0 ? '0倍' : (liveLevel.multiplier % 1 === 0 ? liveLevel.multiplier.toFixed(0) : liveLevel.multiplier.toFixed(1)) + '倍';
    levelCell.innerHTML = '<span class="level-badge ' + levelBadgeClass(liveLevel.name) + '">' + liveLevel.name + '</span> <span class="mult-val">' + levelMultText + '</span>';
    levelCell.setAttribute('data-live', '1');
  }}

  // 4. Q列-近期总涨跌幅：直接用实时估算净值对锚点计算，避免把两段百分比直接相加。
  var qVal = 0;
  if (fd.anchor > 0) {{
    qVal = (estNav - fd.anchor) / fd.anchor * 100;
  }}
  var qCell = row.querySelector('.live-q');
  if (qCell) {{
    var qCls = qVal > 0 ? 'up' : (qVal < 0 ? 'down' : 'flat');
    qCell.innerHTML = '<span class="pct ' + qCls + ' live-est">' + (qVal >= 0 ? '+' : '') + qVal.toFixed(2) + '%</span>';
    qCell.setAttribute('data-live', '1');
  }}

  // 4. R列-补仓份数：qVal 已是百分数（如 -2.35），按整百分点向下取整，信号有效时最低 1 份。
  var shares = 0;
  if (BUY_SIGNALS.indexOf(liveAdvice.action) >= 0) {{
    shares = Math.max(Math.floor(-qVal), 1);
  }}
  var sharesCell = row.querySelector('.live-shares');
  if (sharesCell) {{
    sharesCell.innerHTML = shares > 0
      ? '<span class="shares-val live-est">' + shares + '</span>'
      : '<span class="text-gray live-est">0</span>';
    sharesCell.setAttribute('data-live', shares > 0 ? '2' : '0');
    sharesCell.setAttribute('data-r', shares);
  }}

  // 5. T列-补仓倍数（实时重算）= 当前实时等级倍数 * R。
  var finalMult = liveLevel.multiplier * shares;
  var multCell = row.querySelector('.live-mult');
  if (multCell) {{
    multCell.innerHTML = finalMult > 0
      ? '<span class="guide-mult live-est">' + (finalMult % 1 === 0 ? finalMult.toFixed(0) : finalMult.toFixed(1)) + '</span>'
      : '<span class="text-gray live-est">0</span>';
    multCell.setAttribute('data-live', finalMult > 0 ? '2' : '0');
  }}

  // 6. U列-补仓金额（实时重算）= S * T
  var amount = fd.unit_price * finalMult;
  var amtCell = row.querySelector('.live-amount');
  if (amtCell) {{
    amtCell.innerHTML = amount > 0
      ? '<span class="amt-val live-est">¥' + amount.toLocaleString('en-US', {{maximumFractionDigits:0}}) + '</span>'
      : '<span class="text-gray live-est">¥0</span>';
    amtCell.setAttribute('data-live', amount > 0 ? '2' : '0');
  }}

  // 7. X列-上涨幅度：直接用实时估算净值对最近底谷计算。
  if (fd.trough > 0) {{
    var xVal = (estNav - fd.trough) / fd.trough * 100;
    var xCell = row.querySelector('.live-x');
    if (xCell) {{
      var xCls = xVal > 0 ? 'up' : (xVal < 0 ? 'down' : 'flat');
      xCell.innerHTML = '<span class="pct ' + xCls + ' live-est">' + (xVal >= 0 ? '+' : '') + xVal.toFixed(2) + '%</span>';
      xCell.setAttribute('data-live', '1');
    }}
  }}

  // 存储实时状态到fd对象
  fd._live = true;
  fd._shares = shares;
  fd._amount = amount;
  fd._live_action = liveAdvice.action;
  fd._live_rsi = liveRsi;
}}

// 更新顶部汇总统计
function updateSummary() {{
  var totalShares = 0, totalAmount = 0, buyCount = 0;
  funds.forEach(function(fd) {{
    if (fd._live && fd._shares > 0) {{
      buyCount++;
      totalShares += fd._shares;
      totalAmount += fd._amount;
    }}
  }});
  var statsEl = document.querySelector('.top-bar .stats');
  if (statsEl && totalShares > 0) {{
    statsEl.innerHTML = funds.length + '只基金 — 需补仓<span class="highlight live-est-sum" style="background:#ffd700;color:#1e3a5f;padding:0 4px;border-radius:3px;">' + buyCount + '</span>只 — 共<span class="highlight">' + totalShares + '</span>份 — ¥<span class="highlight">' + totalAmount.toLocaleString('en-US', {{maximumFractionDigits:0}}) + '</span>';
  }}
}}

// 批量读取同站实时估值快照并更新页面。
async function updateAllFunds() {{
  liveSnapshotPromise = null;
  var snapshot = await loadLiveSnapshot();
  var estimates = snapshot && snapshot.estimates ? snapshot.estimates : {{}};
  var liveFunds = snapshot && snapshot.funds ? snapshot.funds : {{}};
  var regimeParams = snapshot && snapshot.market_regime ? snapshot.market_regime.params : null;
  for (var i = 0; i < funds.length; i++) {{
    updateCell(funds[i], estimates[funds[i].code] || null, liveFunds[funds[i].code] || null, regimeParams);
  }}
  updateSseIndex(snapshot);
  updateSummary();
}}

// 时间调速逻辑
function isTradingDay() {{
  var d = new Date();
  var day = d.getDay();
  return day >= 1 && day <= 5;
}}

function getRefreshMode() {{
  // 确认净值只作为静态底稿；交易时段仍允许浏览器临时叠加实时估值。
  if (!isTradingDay()) return 'off';
  var now = new Date();
  var h = now.getHours();
  var m = now.getMinutes();
  var total = h * 60 + m;
  // 9:30-11:30 上午盘中 → 慢速
  if (total >= 570 && total <= 690) return 'slow';
  // 11:30-13:00 午间休市：展示上午收盘的最后一笔估值，不再频繁请求。
  if (total > 690 && total < 780) return 'lunch';
  // 13:00-14:29 下午前段 → 慢速
  if (total >= 780 && total < 870) return 'slow';
  // 14:30-15:00 高频窗口 → 快速
  if (total >= 870 && total <= 900) return 'fast';
  // 15:00后保留当日最后一笔估值，作为盘后复盘参考；快照跨日即自动失效。
  if (total > 900) return 'afterclose';
  return 'off';
}}

// 恢复单行确认净值静态数据（盘后或休市）。
function resetCell(fd) {{
  if (!fd._live) return;
  fd._live = false;
  var row = fd.row;
  var cells = fd.staticCells || {{}};
  var mapping = [
    ['.live-change', 'change'], ['.live-dd', 'dd'], ['.live-rsi', 'rsi'],
    ['.live-action', 'action'], ['.live-level', 'level'], ['.live-q', 'q'],
    ['.live-shares', 'shares'], ['.live-mult', 'mult'], ['.live-amount', 'amount'], ['.live-x', 'x']
  ];
  mapping.forEach(function(item) {{
    var cell = row.querySelector(item[0]);
    if (cell && Object.prototype.hasOwnProperty.call(cells, item[1])) {{
      cell.innerHTML = cells[item[1]];
      cell.setAttribute('data-live', '0');
    }}
  }});
  fd._shares = 0;
  fd._amount = 0;
}}

// 主循环
var countdown = COUNTDOWN_START;

function mainLoop() {{
  var mode = getRefreshMode();
  var statusEl = document.getElementById('live-status');
  var timerEl = document.getElementById('refresh-timer');

  if (mode === 'fast') {{
    if (statusEl) {{ statusEl.textContent = '⚡实时 30s'; statusEl.style.color = '#E67E22'; }}
    updateAllFunds();
    setTimeout(mainLoop, LIVE_INTERVAL);
    countdown = COUNTDOWN_START;
  }} else if (mode === 'slow') {{
    if (statusEl) {{ statusEl.textContent = '🔄盘中 5min'; statusEl.style.color = '#3498db'; }}
    updateAllFunds();
    setTimeout(mainLoop, SLOW_INTERVAL);
    countdown = COUNTDOWN_START;
  }} else if (mode === 'lunch') {{
    if (statusEl) {{ statusEl.textContent = '午间收盘估值'; statusEl.style.color = '#3498db'; }}
    // 用户午间首次打开页面时，拉取一次上午最后报价；已获取后不再重复请求。
    if (!window.__lunchEstimateLoaded) {{
      window.__lunchEstimateLoaded = true;
      updateAllFunds();
    }}
    setTimeout(mainLoop, 60 * 1000);
  }} else if (mode === 'afterclose') {{
    if (statusEl) {{
      statusEl.textContent = '收盘估值参考';
      statusEl.style.color = '#8B7355';
    }}
    // 收盘后首次读取并保留15:00附近最后一笔估值，作为复盘参考；不写回Excel。
    if (!window.__afterCloseEstimateLoaded) {{
      window.__afterCloseEstimateLoaded = true;
      updateAllFunds();
    }}
    if (timerEl) timerEl.textContent = '次日确认';
    setTimeout(mainLoop, 5 * 60 * 1000);
  }} else {{
    if (statusEl) {{
      statusEl.textContent = FINAL_NAV_MODE ? '✓确认净值（休市）' : '⏳休市';
      statusEl.style.color = FINAL_NAV_MODE ? '#27AE60' : '#7f8c8d';
    }}
    // 周末/节假日或无同日快照时恢复Excel确认净值静态底稿。
    funds.forEach(function(fd) {{ resetCell(fd); }});
    countdown--;
    if (timerEl) timerEl.textContent = countdown;
    if (countdown < 0) location.reload();
    setTimeout(mainLoop, 1000);
  }}
}}

// 启动：脚本在 head 中，等待表格 DOM 完整后才读取基金行并开始刷新。
document.addEventListener('DOMContentLoaded', function() {{
  initFunds();
  var timerEl = document.getElementById('refresh-timer');
  if (timerEl) timerEl.textContent = COUNTDOWN_START;
  setTimeout(mainLoop, 1000);
}});
</script>
</head>
<body>
<div class="container">
<div class="mobile-orientation-bar" id="orientation-bar">
  <div class="orientation-copy">
    <div class="orientation-title">完整表格横屏查看</div>
    <div class="orientation-hint" id="orientation-hint">横屏后可同时查看更多列，左侧基金信息会固定。</div>
  </div>
  <div class="orientation-actions">
    <button class="orientation-btn" id="landscape-toggle" type="button">↻ 横屏查看</button>
    <button class="orientation-btn secondary" id="portrait-toggle" type="button">返回竖屏</button>
  </div>
</div>
<p class="rotate-tip" id="rotate-tip">请将手机旋转为横屏；此浏览器不支持自动锁定方向，但横屏后仍会启用紧凑表格和固定基金信息列。</p>

<div class="top-bar">
  <div>
    <span class="main-title">【{date_str}补仓指导】</span>
    <span style="font-size:11px;opacity:0.7;margin-left:12px;"><span id="live-status" style="margin-right:8px;">⏳盘后</span>⏱ <span id="refresh-timer">60</span>秒后刷新</span>
    <span class="stats">{len(funds)}只基金 -- 需补仓<span class="highlight">{need_buy_count}</span>只 -- 共<span class="highlight">{total_shares:,.0f}</span>份 -- ¥<span class="highlight">{total_amount:,.0f}</span></span>
  </div>
  <div style="font-size:12px;opacity:0.85;">数据更新: {update_time}</div>
</div>

<div class="market-bar">
  {market_html}
  <div class="m-item" style="margin-left:auto;"><span class="m-label">说明</span><span class="m-val" style="font-weight:normal;color:#666;">{regime_desc}</span></div>
</div>

<div class="table-wrap">
<table>
<colgroup>
  <col class="col-basic" style="width:75px;background:#E8EDF2">
  <col class="col-basic" style="width:62px;background:#E8EDF2">
  <col class="col-basic" style="width:170px;background:#E8EDF2">
  <col class="col-basic section-divider" style="width:70px;background:#E4EDE6">
  <col class="col-basic" style="width:95px;background:#E4EDE6">
  <col class="col-basic" style="width:100px;background:#E4EDE6">
  <col class="col-basic" style="width:92px;background:#E4EDE6">
  <col class="col-basic" style="width:140px;background:#E4EDE6">
  <col class="col-buy section-divider" style="width:70px;background:#F0E9DD">
  <col class="col-buy" style="width:80px;background:#F0E9DD">
  <col class="col-buy" style="width:55px;background:#F0E9DD">
  <col class="col-buy" style="width:50px;background:#F0E9DD">
  <col class="col-buy" style="width:62px;background:#F0E9DD">
  <col class="col-buy" style="width:75px;background:#F0E9DD">
  <col class="col-tp section-divider" style="width:85px;background:#F9EEF1">
  <col class="col-tp" style="width:75px;background:#F9EEF1">
  <col class="col-tp" style="width:75px;background:#F9EEF1">
  <col class="col-tp" style="width:100px;background:#F9EEF1">
</colgroup>
<thead>
  <tr class="region-row">
    <th colspan="3" class="region-info">基金信息</th>
    <th colspan="5" class="region-indicators">三指标判断详情</th>
    <th colspan="6" class="region-buy">当日补仓指引</th>
    <th colspan="4" class="region-tp">近期止盈参考</th>
  </tr>
  <tr class="header-row">
    <th>类型</th>
    <th>代码</th>
    <th style="text-align:left;">&nbsp;&nbsp;基金名称</th>
    <th>总回撤</th>
    <th>RSI(14)</th>
    <th>PE/PB估值</th>
    <th>操作建议</th>
    <th>倍投等级</th>
    <th>当日涨跌</th>
    <th>近期总涨跌幅</th>
    <th>补仓份数</th>
    <th>补仓单价</th>
    <th>补仓倍数</th>
    <th>补仓金额</th>
    <th>最近底谷日期</th>
    <th>最近底谷净值</th>
    <th>上涨幅度(最近)</th>
    <th>止盈建议</th>
  </tr>
</thead>
<tbody>
{rows_str}
</tbody>
</table>
</div>

<div class="footer">
  <strong>使用说明：</strong><br>
  ① 数据均来自 Excel 实际值（非估算）。F=当日涨跌 Q=近期总涨跌幅 R=补仓份数 S=单价 T=补仓倍数 U=补仓金额。<br>
  ② V/W/X/Y列显示底谷日期/净值/上涨幅度/止盈建议。<br>
  ③ 涨跌幅颜色遵循A股惯例：<span style="color:#e74c3c;font-weight:bold;">红色=涨</span>，<span style="color:#27ae60;font-weight:bold;">绿色=跌</span>。<br>
  ④ 浅黄色背景行 = 操作建议为"关注"的基金。<br>
  ⑤ 系统仅提供数据参谋，投资决策由投资者自行做出。
</div>

</div>
<script>
(function () {{
  var landscapeBtn = document.getElementById('landscape-toggle');
  var portraitBtn = document.getElementById('portrait-toggle');
  var rotateTip = document.getElementById('rotate-tip');
  var hint = document.getElementById('orientation-hint');
  function isLandscape() {{
    return window.matchMedia && window.matchMedia('(orientation: landscape)').matches;
  }}
  function syncOrientationUI() {{
    var landscape = isLandscape();
    document.body.classList.toggle('preview-landscape', landscape);
    if (hint) hint.textContent = landscape
      ? '横屏增强已启用：左侧基金信息固定，可横向浏览完整指标。'
      : '横屏后可同时查看更多列，左侧基金信息会固定。';
    if (rotateTip && landscape) rotateTip.style.display = 'none';
  }}
  async function requestLandscape() {{
    var locked = false;
    try {{
      if (screen.orientation && screen.orientation.lock) {{
        await screen.orientation.lock('landscape');
        locked = true;
      }}
    }} catch (error) {{ locked = false; }}
    if (!locked && rotateTip) rotateTip.style.display = 'block';
    syncOrientationUI();
  }}
  async function requestPortrait() {{
    try {{
      if (screen.orientation && screen.orientation.unlock) screen.orientation.unlock();
    }} catch (error) {{}}
    if (rotateTip) rotateTip.style.display = 'none';
    syncOrientationUI();
  }}
  if (landscapeBtn) landscapeBtn.addEventListener('click', requestLandscape);
  if (portraitBtn) portraitBtn.addEventListener('click', requestPortrait);
  window.addEventListener('orientationchange', syncOrientationUI);
  window.addEventListener('resize', syncOrientationUI);
  syncOrientationUI();
}})();
</script>
</body>
</html>'''
    
    return html

if __name__ == '__main__':
    html = generate()
    with open('金字塔丛林补仓指导图.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'✅ 指导图生成完成！大小: {len(html):,} bytes')
    print(f'   文件: 金字塔丛林补仓指导图.html')
